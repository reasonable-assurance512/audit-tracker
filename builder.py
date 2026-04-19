"""
Audit Resource Tracker — Sprint 1 walking skeleton builder.

Generates a minimal Excel workbook with a single Audit Setup tab.
Full fidelity to the v4 workbook specification is deferred to Sprint 2.
"""

from datetime import date, timedelta
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


DARK_BLUE = "1F3864"
PURPLE = "7030A0"
LIGHT_BLUE = "DCE6F1"
LIGHT_GRN = "EBF1DE"
GRAY_MD = "D9D9D9"
GRAY_LT = "F2F2F2"
BLUE_IN = "0000CD"
WHITE = "FFFFFF"
BLACK = "000000"


def _font(bold=False, color=BLACK, size=10, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)


def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, end_color=hex_color)


def _align(horizontal="left", vertical="center", wrap=False):
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)


def _border():
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def build_workbook(kickoff_date, planning_weeks, fieldwork_weeks, reporting_weeks):
    """
    Build a minimal Audit Resource Tracker workbook.

    Args:
        kickoff_date: date object for the audit kickoff
        planning_weeks: int, number of planning phase weeks
        fieldwork_weeks: int, number of fieldwork phase weeks
        reporting_weeks: int, number of reporting phase weeks

    Returns:
        BytesIO containing the generated .xlsx file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Setup"
    ws.sheet_properties.tabColor = PURPLE

    for col_letter, width in [("A", 36), ("B", 16), ("C", 16), ("D", 16), ("E", 46)]:
        ws.column_dimensions[col_letter].width = width

    ws.merge_cells("A1:E1")
    title = ws["A1"]
    title.value = "Audit Configuration & Milestone Calculator | Sprint 1 Preview"
    title.font = _font(bold=True, color=WHITE, size=13)
    title.fill = _fill(PURPLE)
    title.alignment = _align("center")
    ws.row_dimensions[1].height = 26

    ws.row_dimensions[2].height = 6

    ws.merge_cells("A3:E3")
    hdr = ws["A3"]
    hdr.value = "ANCHOR DATES"
    hdr.font = _font(bold=True, color=WHITE, size=10)
    hdr.fill = _fill(PURPLE)
    hdr.alignment = _align("left")
    ws.row_dimensions[3].height = 18

    for col_idx, label in enumerate(["Field", "Value", "Day of Week", "", "Notes"], 1):
        cell = ws.cell(row=4, column=col_idx, value=label)
        cell.font = _font(bold=True, size=9)
        cell.fill = _fill(GRAY_MD)
        cell.alignment = _align("center")
        cell.border = _border()

    ws.cell(row=5, column=1, value="Kickoff / Project Launch Date").font = _font(bold=True, size=10)
    ws.cell(row=5, column=1).border = _border()

    kickoff_cell = ws.cell(row=5, column=2, value=kickoff_date)
    kickoff_cell.font = _font(bold=True, color=BLUE_IN, size=11)
    kickoff_cell.fill = _fill(LIGHT_BLUE)
    kickoff_cell.number_format = "MM/DD/YYYY"
    kickoff_cell.alignment = _align("center")
    kickoff_cell.border = _border()

    ws.cell(row=5, column=3, value=kickoff_date.strftime("%A")).alignment = _align("center")
    ws.cell(row=5, column=3).border = _border()

    total_weeks = planning_weeks + fieldwork_weeks + reporting_weeks
    kickoff_monday = kickoff_date - timedelta(days=kickoff_date.weekday())
    release_date = kickoff_monday + timedelta(weeks=total_weeks)

    ws.cell(row=6, column=1, value="Report Release Date (calculated)").font = _font(bold=True, size=10)
    ws.cell(row=6, column=1).border = _border()

    release_cell = ws.cell(row=6, column=2, value=release_date)
    release_cell.font = _font(bold=True, size=11)
    release_cell.fill = _fill(LIGHT_GRN)
    release_cell.number_format = "MM/DD/YYYY"
    release_cell.alignment = _align("center")
    release_cell.border = _border()

    ws.cell(row=6, column=3, value=release_date.strftime("%A")).alignment = _align("center")
    ws.cell(row=6, column=3).border = _border()

    ws.row_dimensions[7].height = 8

    ws.merge_cells("A8:E8")
    phase_hdr = ws["A8"]
    phase_hdr.value = "PHASE CONFIGURATION"
    phase_hdr.font = _font(bold=True, color=WHITE, size=10)
    phase_hdr.fill = _fill(DARK_BLUE)
    phase_hdr.alignment = _align("left")
    ws.row_dimensions[8].height = 18

    for col_idx, label in enumerate(["Phase", "Weeks", "Start (Mon)", "End (Fri)", "Notes"], 1):
        cell = ws.cell(row=9, column=col_idx, value=label)
        cell.font = _font(bold=True, size=9)
        cell.fill = _fill(GRAY_MD)
        cell.alignment = _align("center")
        cell.border = _border()

    plan_start = kickoff_monday
    plan_end = plan_start + timedelta(weeks=planning_weeks) - timedelta(days=3)
    field_start = plan_end + timedelta(days=3)
    field_end = field_start + timedelta(weeks=fieldwork_weeks) - timedelta(days=3)
    rep_start = field_end + timedelta(days=3)
    rep_end = rep_start + timedelta(weeks=reporting_weeks) - timedelta(days=3)

    phases = [
        ("PLANNING", planning_weeks, plan_start, plan_end),
        ("FIELDWORK", fieldwork_weeks, field_start, field_end),
        ("REPORTING", reporting_weeks, rep_start, rep_end),
    ]

    for row_idx, (phase_name, weeks, start, end) in enumerate(phases, 10):
        ws.cell(row=row_idx, column=1, value=phase_name).font = _font(bold=True, size=10)
        ws.cell(row=row_idx, column=1).border = _border()

        weeks_cell = ws.cell(row=row_idx, column=2, value=weeks)
        weeks_cell.font = _font(bold=True, size=10)
        weeks_cell.fill = _fill(LIGHT_BLUE)
        weeks_cell.alignment = _align("center")
        weeks_cell.border = _border()
        weeks_cell.number_format = '0" wks"'

        start_cell = ws.cell(row=row_idx, column=3, value=start)
        start_cell.number_format = "MM/DD/YY"
        start_cell.alignment = _align("center")
        start_cell.border = _border()

        end_cell = ws.cell(row=row_idx, column=4, value=end)
        end_cell.number_format = "MM/DD/YY"
        end_cell.alignment = _align("center")
        end_cell.border = _border()

        ws.cell(row=row_idx, column=5, value="").border = _border()

    ws.cell(row=13, column=1, value="TOTAL AUDIT WEEKS").font = _font(bold=True, size=10)
    ws.cell(row=13, column=1).fill = _fill(GRAY_LT)
    ws.cell(row=13, column=1).border = _border()

    total_cell = ws.cell(row=13, column=2, value=total_weeks)
    total_cell.font = _font(bold=True, size=12)
    total_cell.fill = _fill(GRAY_MD)
    total_cell.alignment = _align("center")
    total_cell.border = _border()
    total_cell.number_format = '0" wks"'

    wb.properties.creator = "Audit Resource Tracker"
    wb.properties.lastModifiedBy = ""
    wb.properties.title = ""
    wb.properties.subject = ""
    wb.properties.description = ""

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


if __name__ == "__main__":
    test_kickoff = date(2026, 5, 4)
    output = build_workbook(test_kickoff, 4, 16, 4)
    with open("test_output.xlsx", "wb") as f:
        f.write(output.read())
    print("Test workbook written to test_output.xlsx")
