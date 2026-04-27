"""
Parity tests comparing the modularized builder output against the
v4 golden file. A cell-by-cell comparison of values, formulas, and
key formatting properties for each tab that has been extracted so far.

Run with: pytest tests/test_parity.py -v
Or directly: python tests/test_parity.py
"""

import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from builder.config import AuditConfig
from builder.constants import RES_TABS
from builder.holidays_tab import build_holidays_tab
from builder.resource_tab import build_all_resource_tabs
from builder.setup_tab import build_setup_tab


GOLDEN_FILE_PATH = PROJECT_ROOT / "tests" / "golden_files" / "v4_reference.xlsx"


def generate_full_workbook():
    """
    Generate a workbook with all tabs extracted so far.
    Currently: Holidays & Skeleton, Audit Setup, 9 Resource tabs.

    Uses a default-valued AuditConfig so the output matches the
    v4 reference golden file. Different config values would produce
    different output and fail parity (which is correct — parity is
    only meaningful for the default-configured case).
    """
    wb = Workbook()
    config = AuditConfig()
    holidays_info = build_holidays_tab(wb)
    build_setup_tab(
        wb,
        config=config,
        closed_range=holidays_info["closed_range"],
        skeleton_range=holidays_info["skeleton_range"],
    )
    build_all_resource_tabs(
        wb,
        config=config,
        closed_range=holidays_info["closed_range"],
        skeleton_range=holidays_info["skeleton_range"],
    )
    return wb


def normalize_value(value):
    """Normalize xlsx save/load artifacts."""
    if isinstance(value, datetime) and value.time().hour == 0 and value.time().minute == 0:
        return value.date()
    if value == "":
        return None
    return value


def normalize_wrap_text(value):
    """Treat wrap_text=False and wrap_text=None as equivalent."""
    if value is None:
        return False
    return value


def get_cell_snapshot(cell):
    """Extract comparable properties from a cell."""
    return {
        "value": normalize_value(cell.value),
        "number_format": cell.number_format,
        "font_bold": cell.font.bold,
        "font_italic": cell.font.italic,
        "font_color": cell.font.color.rgb if cell.font.color else None,
        "font_size": cell.font.size,
        "fill_color": (
            cell.fill.start_color.rgb
            if cell.fill and cell.fill.start_color
            else None
        ),
        "horizontal": cell.alignment.horizontal,
        "vertical": cell.alignment.vertical,
        "wrap_text": normalize_wrap_text(cell.alignment.wrap_text),
    }


def compare_tabs(new_ws, golden_ws, tab_name):
    """Cell-by-cell comparison. Returns list of differences."""
    differences = []

    max_row = max(new_ws.max_row, golden_ws.max_row)
    max_col = max(new_ws.max_column, golden_ws.max_column)

    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            new_cell = new_ws.cell(row=row, column=col)
            golden_cell = golden_ws.cell(row=row, column=col)

            if new_cell.value is None and golden_cell.value is None:
                continue
            if new_cell.value == "" and golden_cell.value is None:
                continue
            if new_cell.value is None and golden_cell.value == "":
                continue

            new_snap = get_cell_snapshot(new_cell)
            golden_snap = get_cell_snapshot(golden_cell)

            if new_snap != golden_snap:
                differences.append({
                    "tab": tab_name,
                    "cell": f"{new_cell.coordinate}",
                    "new": new_snap,
                    "golden": golden_snap,
                })

    return differences


def run_parity_check(tab_name):
    """Generate a workbook, load golden, compare specified tab."""
    new_wb = generate_full_workbook()
    new_ws = new_wb[tab_name]

    golden_wb = load_workbook(GOLDEN_FILE_PATH)
    golden_ws = golden_wb[tab_name]

    return compare_tabs(new_ws, golden_ws, tab_name)


def report_and_assert(tab_name, differences):
    """Print differences (first 10) and assert zero if any."""
    if differences:
        print(f"\n{tab_name}: {len(differences)} differences found")
        for diff in differences[:10]:
            print(f"  Cell {diff['cell']}:")
            print(f"    new:    {diff['new']}")
            print(f"    golden: {diff['golden']}")
        if len(differences) > 10:
            print(f"  ... and {len(differences) - 10} more")
        raise AssertionError(
            f"{tab_name} has {len(differences)} cell differences"
        )
    print(f"{tab_name}: PARITY VERIFIED (no differences)")


def test_holidays_tab_parity():
    """Verify the modularized holidays tab matches the golden file."""
    differences = run_parity_check("Holidays & Skeleton")
    report_and_assert("Holidays & Skeleton", differences)


def test_setup_tab_parity():
    """Verify the modularized setup tab matches the golden file."""
    differences = run_parity_check("Audit Setup")
    report_and_assert("Audit Setup", differences)


def test_resource_tabs_parity():
    """Verify all 9 modularized resource tabs match the golden file."""
    tab_diff_counts = {}
    for tab_name in RES_TABS:
        differences = run_parity_check(tab_name)
        if differences:
            tab_diff_counts[tab_name] = differences

    if tab_diff_counts:
        first_tab = list(tab_diff_counts.keys())[0]
        diffs = tab_diff_counts[first_tab]
        print(f"\n{first_tab}: {len(diffs)} differences found (first tab with diffs)")
        for diff in diffs[:10]:
            print(f"  Cell {diff['cell']}:")
            print(f"    new:    {diff['new']}")
            print(f"    golden: {diff['golden']}")
        if len(diffs) > 10:
            print(f"  ... and {len(diffs) - 10} more in {first_tab}")
        other_tabs = [
            f"{t}: {len(d)}" for t, d in list(tab_diff_counts.items())[1:]
        ]
        if other_tabs:
            print(f"  Other tabs with differences: {', '.join(other_tabs)}")
        total = sum(len(d) for d in tab_diff_counts.values())
        raise AssertionError(
            f"Resource tabs have {total} total cell differences "
            f"across {len(tab_diff_counts)} tabs"
        )

    print(f"All 9 Resource tabs: PARITY VERIFIED (no differences)")


# ─────────────────────────────────────────────────────────────────
# Deferred parity coverage
# ─────────────────────────────────────────────────────────────────
# The Master Budget by Date and Budget by Task tabs are not yet
# covered by parity tests. Per Option D (Sprint 2 Step 5 design
# decision), these tabs are extracted into the modular package and
# verified visually but not by automated cell-by-cell comparison.
#
# Tracked as backlog item F-10 for a future sprint.
# ─────────────────────────────────────────────────────────────────


if __name__ == "__main__":
    test_holidays_tab_parity()
    test_setup_tab_parity()
    test_resource_tabs_parity()
