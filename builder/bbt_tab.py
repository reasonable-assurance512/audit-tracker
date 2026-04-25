"""
Builder for the Budget by Task tab.

Cost-tracking tab organized by audit task and resource. Combines
auto-populated rows (sourced from MBDD subtotals) with manual entry
rows (Fieldwork tasks, Coordination codes). Includes a Fieldwork
Ceiling Validation row that flags over-allocation per resource.

Layout:
  Row 1: Title
  Row 2: Spacer
  Row 3: Note bar
  Row 4: Column headers
  Row 5+: Four sections (Planning, Fieldwork, Reporting, Cross-Cutting)
  Row 36: Grand Total
"""

from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font

from .constants import (
    AM_CLR,
    BBT_CC_AM,
    BBT_CC_DIV,
    BBT_CC_MGMT,
    BBT_CC_MTG,
    BBT_CC_QC,
    BBT_CC_RE,
    BBT_CC_TOT,
    BBT_CC_WRAP,
    BBT_FIELD_DIV,
    BBT_FIELD_E,
    BBT_FIELD_S,
    BBT_FIELD_TOT,
    BBT_FIELD_VAL,
    BBT_GRAND,
    BBT_PLAN_DIV,
    BBT_PLAN_ROW,
    BBT_PLAN_TOT,
    BBT_REP_DIV,
    BBT_REP1,
    BBT_REP2,
    BBT_REP_TOT,
    BLACK,
    BLUE_IN,
    DARK_BLUE,
    DARK_RED,
    FIELD_HDR,
    GRAY_LT,
    GRAY_MD,
    LIGHT_BLUE,
    LIGHT_PURPLE,
    LIGHT_RED,
    MBDD_FIELD_TOT,
    MBDD_GRAND,
    MBDD_PLAN_TOT,
    MBDD_REP_TOT,
    MBDD_RES_COLS,
    MED_BLUE,
    PLAN_BG,
    PLAN_HDR,
    QC_CLR,
    REP_BG,
    REP_HDR,
    RES_NAMES,
    TEAL,
    WARN_RED,
    WHITE,
    column_letter,
    make_align,
    make_border,
    make_fill,
    make_font,
)


BBT_COLS = {
    "A": 28, "B": 18, "C": 14, "D": 14, "E": 14,
    "F": 14, "G": 14, "H": 14, "I": 14, "J": 14, "K": 14,
}

# Resource categories used by various sections
STAFF_RES = RES_NAMES[:6]   # PM, Asst PM, Auditors 1-4
OH_RES = RES_NAMES[6:]      # AM, QC, RE


def _mbdd_col(name):
    """Return the column letter on MBDD for a given resource name."""
    return column_letter(MBDD_RES_COLS[name])


def _write_title_row(ws):
    """Row 1: title."""
    ws.merge_cells("A1:K1")
    title = ws["A1"]
    title.value = "Budget by Task | v4"
    title.font = make_font(bold=True, color=WHITE, size=13)
    title.fill = make_fill(DARK_RED)
    title.alignment = make_align("center")
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 8


def _write_note_row(ws):
    """Row 3 currently doubles as the section divider for Planning.
    Planning section uses row 3 as its divider per BBT_PLAN_DIV.
    No separate note row in this layout."""
    pass


def _write_column_headers(ws, header_row):
    """Column header row (row 4 = the row before Planning divider conceptually)."""
    ws.row_dimensions[header_row].height = 34

    color_map = {
        "AM": AM_CLR,
        "QC": QC_CLR,
        "RE": TEAL,
    }
    headers = [
        (1, "Task Name / Code", DARK_BLUE, WHITE),
        (2, "Budget Est.\n(Total)", DARK_RED, WHITE),
    ]
    for i, name in enumerate(RES_NAMES):
        bg = color_map.get(name, MED_BLUE)
        headers.append((3 + i, name, bg, WHITE))

    for col_idx, label, bg, fg in headers:
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = make_font(bold=True, color=fg, size=9)
        cell.fill = make_fill(bg)
        cell.alignment = make_align("center", wrap=True)
        cell.border = make_border()


def _write_section_divider(ws, row, label, bg_color):
    """Section header bar spanning A-K."""
    ws.row_dimensions[row].height = 22
    ws.merge_cells(f"A{row}:K{row}")
    cell = ws[f"A{row}"]
    cell.value = f" \u25b8 {label}"
    cell.font = make_font(bold=True, color=WHITE, size=11)
    cell.fill = make_fill(bg_color)
    cell.alignment = make_align("left")


def _write_auto_row(ws, row, label, resource_formulas, bg_color=WHITE):
    """
    Auto-populated row.
    label is the task name (column A).
    resource_formulas is a dict mapping resource name to formula string
    (or empty string for resources that don't apply).
    Budget Est. column (B) is auto-calculated as SUM(C:K) for that row.
    """
    ws.row_dimensions[row].height = 17

    name_cell = ws.cell(row=row, column=1, value=label)
    name_cell.font = make_font(size=10)
    name_cell.fill = make_fill(bg_color)
    name_cell.border = make_border()
    name_cell.alignment = make_align()

    for i, name in enumerate(RES_NAMES):
        col_idx = 3 + i
        formula = resource_formulas.get(name, "")
        cell = ws.cell(row=row, column=col_idx, value=formula if formula else 0)
        cell.font = make_font(size=9, color="595959" if not formula else BLACK)
        cell.fill = make_fill(bg_color)
        cell.number_format = '#,##0;-#,##0;"-"'
        cell.alignment = make_align("center")
        cell.border = make_border()

    budget_cell = ws.cell(row=row, column=2, value=f"=SUM(C{row}:K{row})")
    budget_cell.font = make_font(bold=True, size=10)
    budget_cell.fill = make_fill(bg_color)
    budget_cell.number_format = "#,##0"
    budget_cell.alignment = make_align("center")
    budget_cell.border = make_border()


def _write_manual_row(ws, row, label_or_blank=""):
    """
    Manual entry row (Fieldwork tasks or coordination codes).
    All cells start blank/zero with light-blue fill to indicate input.
    Budget Est. column auto-sums the resource columns.
    """
    ws.row_dimensions[row].height = 17

    name_cell = ws.cell(row=row, column=1, value=label_or_blank)
    name_cell.font = make_font(size=9, color=BLUE_IN, italic=True)
    name_cell.fill = make_fill(LIGHT_BLUE)
    name_cell.border = make_border()
    name_cell.alignment = make_align()

    budget_cell = ws.cell(row=row, column=2, value=f"=SUM(C{row}:K{row})")
    budget_cell.font = make_font(bold=True, size=9)
    budget_cell.fill = make_fill(LIGHT_BLUE)
    budget_cell.number_format = '#,##0;-#,##0;"-"'
    budget_cell.alignment = make_align("center")
    budget_cell.border = make_border()

    for col_idx in range(3, 12):
        cell = ws.cell(row=row, column=col_idx, value=0)
        cell.font = make_font(color=BLUE_IN, size=9)
        cell.fill = make_fill(LIGHT_BLUE)
        cell.number_format = '#,##0;-#,##0;"-"'
        cell.alignment = make_align("center")
        cell.border = make_border()


def _write_total_row(ws, row, label, data_rows, bg_color):
    """
    Section total row.
    data_rows is a list of row numbers being summed.
    Budget Est. column sums across all 9 resource columns of the totals.
    """
    ws.row_dimensions[row].height = 20
    label_cell = ws.cell(row=row, column=1, value=f" {label}")
    label_cell.font = make_font(bold=True, color=WHITE, size=10)
    label_cell.fill = make_fill(bg_color)
    label_cell.border = make_border()
    label_cell.alignment = make_align()

    for col_idx in range(2, 12):
        col_letter = column_letter(col_idx)
        sum_expr = "+".join([f"{col_letter}{r}" for r in data_rows])
        cell = ws.cell(row=row, column=col_idx, value=f"={sum_expr}")
        cell.font = make_font(bold=True, color=WHITE, size=10)
        cell.fill = make_fill(bg_color)
        cell.number_format = "#,##0"
        cell.alignment = make_align("center")
        cell.border = make_border()


def _write_fieldwork_validation_row(ws):
    """
    Row that validates Fieldwork allocations against MBDD ceilings.
    For each resource column, formula compares SUM of Fieldwork rows
    to the corresponding MBDD Fieldwork subtotal.
    Result text: "OVER" if exceeded, "OK" if not.
    Conditional formatting applies red fill to OVER cells.
    """
    ws.row_dimensions[BBT_FIELD_VAL].height = 20

    label = ws.cell(
        row=BBT_FIELD_VAL,
        column=1,
        value=" \u26a0 Fieldwork Ceiling Check (allocated vs. MBDD)",
    )
    label.font = make_font(bold=True, size=9, color=DARK_RED)
    label.fill = make_fill(GRAY_LT)
    label.border = make_border()
    label.alignment = make_align()

    ws.cell(row=BBT_FIELD_VAL, column=2).fill = make_fill(GRAY_LT)
    ws.cell(row=BBT_FIELD_VAL, column=2).border = make_border()

    for i, name in enumerate(RES_NAMES):
        col_idx = 3 + i
        col_letter = column_letter(col_idx)
        mbdd_col = _mbdd_col(name)

        validation_formula = (
            f"=IF(SUM({col_letter}{BBT_FIELD_S}:{col_letter}{BBT_FIELD_E})"
            f">'Master Budget by Date'!{mbdd_col}{MBDD_FIELD_TOT},"
            f'"\u26a0 OVER","\u2713 OK")'
        )
        cell = ws.cell(row=BBT_FIELD_VAL, column=col_idx, value=validation_formula)
        cell.font = make_font(bold=True, size=9)
        cell.fill = make_fill(GRAY_LT)
        cell.alignment = make_align("center")
        cell.border = make_border()

        # CF: red if over
        ws.conditional_formatting.add(
            f"{col_letter}{BBT_FIELD_VAL}",
            FormulaRule(
                formula=[
                    f"=SUM({col_letter}{BBT_FIELD_S}:{col_letter}{BBT_FIELD_E})"
                    f">'Master Budget by Date'!{mbdd_col}{MBDD_FIELD_TOT}"
                ],
                fill=make_fill(LIGHT_RED),
                font=Font(name="Arial", bold=True, color=WARN_RED, size=9),
            ),
        )


def _write_grand_total(ws):
    """Grand total row summing all four section totals."""
    ws.row_dimensions[BBT_GRAND].height = 24
    label = ws.cell(
        row=BBT_GRAND,
        column=1,
        value="GRAND TOTAL - ALL PHASES & CROSS-CUTTING",
    )
    label.font = make_font(bold=True, color=WHITE, size=11)
    label.fill = make_fill(DARK_RED)
    label.border = make_border()
    label.alignment = make_align()

    for col_idx in range(2, 12):
        col_letter = column_letter(col_idx)
        formula = (
            f"={col_letter}{BBT_PLAN_TOT}+{col_letter}{BBT_FIELD_TOT}"
            f"+{col_letter}{BBT_REP_TOT}+{col_letter}{BBT_CC_TOT}"
        )
        cell = ws.cell(row=BBT_GRAND, column=col_idx, value=formula)
        cell.font = make_font(bold=True, color=WHITE, size=12)
        cell.fill = make_fill(DARK_RED)
        cell.number_format = "#,##0"
        cell.alignment = make_align("center")
        cell.border = make_border()


def build_bbt_tab(wb, closed_range, skeleton_range):
    """
    Build the Budget by Task tab.

    Args:
        wb: openpyxl Workbook (must already have MBDD tab)
        closed_range: (accepted for interface consistency, unused)
        skeleton_range: (accepted for interface consistency, unused)

    Returns:
        dict with tab metadata.
    """
    ws = wb.create_sheet("Budget by Task")
    ws.sheet_properties.tabColor = "C00000"

    for col_letter, width in BBT_COLS.items():
        ws.column_dimensions[col_letter].width = width

    _write_title_row(ws)
    _write_column_headers(ws, header_row=2)

    # ── Section 1: Planning ─────────────────────────────────
    _write_section_divider(ws, BBT_PLAN_DIV, "PLANNING", PLAN_HDR)
    plan_formulas = {
        name: f"='Master Budget by Date'!{_mbdd_col(name)}{MBDD_PLAN_TOT}"
        for name in STAFF_RES
    }
    for name in OH_RES:
        plan_formulas[name] = ""
    _write_auto_row(ws, BBT_PLAN_ROW, "100 - Overall Planning", plan_formulas, PLAN_BG)
    _write_total_row(ws, BBT_PLAN_TOT, "Planning Total", [BBT_PLAN_ROW], PLAN_HDR)

    # ── Section 2: Fieldwork ─────────────────────────────────
    _write_section_divider(
        ws, BBT_FIELD_DIV,
        "FIELDWORK (enter tasks below - up to 15 rows)",
        FIELD_HDR,
    )
    for row in range(BBT_FIELD_S, BBT_FIELD_E + 1):
        _write_manual_row(ws, row)
    _write_total_row(
        ws, BBT_FIELD_TOT, "Fieldwork Total",
        list(range(BBT_FIELD_S, BBT_FIELD_E + 1)), FIELD_HDR,
    )
    _write_fieldwork_validation_row(ws)

    # ── Section 3: Reporting ─────────────────────────────────
    _write_section_divider(ws, BBT_REP_DIV, "REPORTING", REP_HDR)
    rep_formulas = {
        name: f"='Master Budget by Date'!{_mbdd_col(name)}{MBDD_REP_TOT}"
        for name in STAFF_RES
    }
    for name in OH_RES:
        rep_formulas[name] = ""
    _write_auto_row(ws, BBT_REP1, "300 - Reporting", rep_formulas, REP_BG)

    # 377 - Technical Writing (manual entry on this row)
    _write_manual_row(ws, BBT_REP2, label_or_blank="377 - Technical Writing (manual)")

    _write_total_row(
        ws, BBT_REP_TOT, "Reporting Total",
        [BBT_REP1, BBT_REP2], REP_HDR,
    )

    # ── Section 4: Cross-Cutting ─────────────────────────────
    _write_section_divider(
        ws, BBT_CC_DIV,
        "CROSS-CUTTING (overhead roles + coordination)",
        DARK_BLUE,
    )

    # 032 - Audit Manager: pulls from MBDD AM column grand total
    am_formulas = {
        name: (f"='Master Budget by Date'!{_mbdd_col('AM')}{MBDD_GRAND}" if name == "AM" else "")
        for name in RES_NAMES
    }
    _write_auto_row(ws, BBT_CC_AM, "032 - Audit Manager", am_formulas, LIGHT_PURPLE)

    # 001 - Quality Control Review: pulls from MBDD QC column grand total
    qc_formulas = {
        name: (f"='Master Budget by Date'!{_mbdd_col('QC')}{MBDD_GRAND}" if name == "QC" else "")
        for name in RES_NAMES
    }
    _write_auto_row(ws, BBT_CC_QC, "001 - Quality Control Review", qc_formulas, LIGHT_PURPLE)

    # 014 - Report Editing: pulls from MBDD RE column grand total
    re_formulas = {
        name: (f"='Master Budget by Date'!{_mbdd_col('RE')}{MBDD_GRAND}" if name == "RE" else "")
        for name in RES_NAMES
    }
    _write_auto_row(ws, BBT_CC_RE, "014 - Report Editing", re_formulas, LIGHT_PURPLE)

    # 004, 060, 070 - Coordination codes (manual entry)
    _write_manual_row(ws, BBT_CC_MTG, label_or_blank="004 - Project Team Meetings")
    _write_manual_row(ws, BBT_CC_WRAP, label_or_blank="060 - Project Wrap-Up")
    _write_manual_row(ws, BBT_CC_MGMT, label_or_blank="070 - Project Management")

    _write_total_row(
        ws, BBT_CC_TOT, "Cross-Cutting Total",
        [BBT_CC_AM, BBT_CC_QC, BBT_CC_RE, BBT_CC_MTG, BBT_CC_WRAP, BBT_CC_MGMT],
        DARK_BLUE,
    )

    # ── Grand Total ─────────────────────────────────────────
    _write_grand_total(ws)

    ws.freeze_panes = "A5"

    return {"tab_name": "Budget by Task"}
