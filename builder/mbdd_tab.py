"""
Builder for the Master Budget by Date (MBDD) tab.

Read-only consolidation tab that aggregates hours from all 9 resource
tabs by week. Cross-tab formulas reference the H column (Total Available
Hours) of each resource tab. Includes phase subtotals, grand total,
running cumulative total, and milestone/warning text.

Row layout uses an offset relative to resource tabs:
  Planning:   MBDD row r corresponds to resource row r+2
  Fieldwork:  MBDD row r corresponds to resource row r+1
  Reporting:  MBDD row r corresponds to resource row r (same)

This offset arises because MBDD's section dividers and resource tabs'
section dividers fall in slightly different rows due to MBDD having
an extra protection-note row near the top.
"""

from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font

from .config import AuditConfig

from .constants import (
    AM_CLR,
    AS_FIELD,
    AS_PLAN,
    AS_REP,
    AS_TOT,
    BLACK,
    DARK_BLUE,
    DARK_GRN,
    DEFAULT_FIELD,
    DEFAULT_PLAN,
    DEFAULT_REP,
    FIELD_BG,
    FIELD_HDR,
    GRAY_LT,
    LIGHT_GRN,
    MBDD_FIELD_DIV,
    MBDD_FIELD_E,
    MBDD_FIELD_S,
    MBDD_FIELD_TOT,
    MBDD_GRAND,
    MBDD_PLAN_DIV,
    MBDD_PLAN_E,
    MBDD_PLAN_S,
    MBDD_PLAN_TOT,
    MBDD_REP_DIV,
    MBDD_REP_E,
    MBDD_REP_S,
    MBDD_REP_TOT,
    MED_BLUE,
    MILESTONE_FG,
    MS_BG,
    MS_ROWS,
    PLAN_BG,
    PLAN_HDR,
    PURPLE,
    QC_CLR,
    REP_BG,
    REP_HDR,
    RES_NAMES,
    RES_TABS,
    TEAL,
    WARN_BG,
    WARN1_BG,
    WARN2_BG,
    WARN_RED,
    WHITE,
    column_letter,
    make_align,
    make_border,
    make_fill,
    make_font,
)


MBDD_COLS = {
    "A": 14, "B": 11, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14,
    "H": 14, "I": 14, "J": 14, "K": 14, "L": 14, "M": 14, "N": 36,
}


def _milestone_warning_formula(date_cell_ref):
    """
    Build the column N formula that checks 8 major milestones against
    the row's date and produces three states per milestone:
      - this week: flag symbol + name
      - one week away: 1-week warning
      - two weeks away (selected milestones only): 2-week warning
    """
    milestone_checks = [
        (MS_ROWS["kickoff"], "\u2691 KICKOFF"),
        (MS_ROWS["end_plan"], "\u2691 END OF PLANNING"),
        (MS_ROWS["outline"], "\u2691 OUTLINE TO EXEC"),
        (MS_ROWS["ontarget"], "\u2691 ON-TARGET MEETING"),
        (MS_ROWS["draft"], "\u2691 DRAFT TO EXEC"),
        (MS_ROWS["exit"], "\u2691 EXIT CONFERENCE"),
        (MS_ROWS["mgmt"], "\u2691 MGMT RESPONSES DUE"),
        (MS_ROWS["release"], "\u2691 REPORT RELEASE"),
    ]
    week1_warnings = [
        (MS_ROWS["kickoff"], "\u26a0 NEXT WEEK: KICKOFF"),
        (MS_ROWS["end_plan"], "\u26a0 NEXT WEEK: END PLANNING"),
        (MS_ROWS["outline"], "\u26a0 NEXT WEEK: OUTLINE TO EXEC"),
        (MS_ROWS["ontarget"], "\u26a0 NEXT WEEK: ON-TARGET"),
        (MS_ROWS["draft"], "\u26a0 NEXT WEEK: DRAFT TO EXEC"),
        (MS_ROWS["exit"], "\u26a0 NEXT WEEK: EXIT CONFERENCE"),
        (MS_ROWS["mgmt"], "\u26a0 NEXT WEEK: MGMT RESPONSES"),
        (MS_ROWS["release"], "\u26a0 NEXT WEEK: RELEASE"),
    ]
    week2_warnings = [
        (MS_ROWS["ontarget"], "\u26a0 2 WEEKS: ON-TARGET - PREPARE"),
        (MS_ROWS["draft"], "\u26a0 2 WEEKS: DRAFT TO EXEC - PREPARE"),
        (MS_ROWS["release"], "\u26a0 2 WEEKS: RELEASE - PREPARE"),
    ]

    parts = []
    for r, label in milestone_checks:
        parts.append(
            f"IF(AND('Audit Setup'!$C${r}>={date_cell_ref},"
            f"'Audit Setup'!$C${r}<={date_cell_ref}+4),\"{label} \",\"\")"
        )
    for r, label in week1_warnings:
        parts.append(
            f"IF(AND('Audit Setup'!$C${r}-7>={date_cell_ref},"
            f"'Audit Setup'!$C${r}-7<={date_cell_ref}+4),\"{label} \",\"\")"
        )
    for r, label in week2_warnings:
        parts.append(
            f"IF(AND('Audit Setup'!$C${r}-14>={date_cell_ref},"
            f"'Audit Setup'!$C${r}-14<={date_cell_ref}+4),\"{label} \",\"\")"
        )

    inner = ",".join(parts)
    return (
        f'=IF(NOT(ISNUMBER({date_cell_ref})),"",'
        f'TRIM(CONCATENATE({inner})))'
    )


def _write_title_block(ws):
    """Title row plus protection note row."""
    ws.merge_cells("A1:N1")
    title = ws["A1"]
    title.value = "Master Budget by Date | READ-ONLY | v4"
    title.font = make_font(bold=True, color=WHITE, size=13)
    title.fill = make_fill(DARK_BLUE)
    title.alignment = make_align("center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:N2")
    note = ws["A2"]
    note.value = (
        "This tab is read-only and formula-driven. To change hours or leave, "
        "go to the relevant Resource tab."
    )
    note.font = make_font(italic=True, bold=True, size=9, color=WARN_RED)
    note.fill = make_fill(WARN_BG)
    note.alignment = make_align("center", wrap=True)
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 6


def _write_column_headers(ws):
    """Row 4 — 14 column headers."""
    ws.row_dimensions[4].height = 34
    headers = [
        (1, "Week Of\n(Monday)", DARK_BLUE, WHITE),
        (2, "Phase\nLabel", DARK_BLUE, WHITE),
    ]
    color_map = {
        "PM": MED_BLUE,
        "Asst PM": MED_BLUE,
        "AM": AM_CLR,
        "QC": QC_CLR,
        "RE": TEAL,
    }
    for i, name in enumerate(RES_NAMES):
        bg = color_map.get(name, FIELD_HDR)
        headers.append((3 + i, name, bg, WHITE))
    headers.extend([
        (12, "Weekly\nTotal", DARK_GRN, WHITE),
        (13, "Total\nto Date", DARK_GRN, WHITE),
        (14, "Milestones & Warnings\n(auto)", PURPLE, WHITE),
    ])
    for col_idx, label, bg, fg in headers:
        cell = ws.cell(row=4, column=col_idx, value=label)
        cell.font = make_font(bold=True, color=fg, size=9)
        cell.fill = make_fill(bg)
        cell.alignment = make_align("center", wrap=True)
        cell.border = make_border()


def _write_phase_divider(ws, row, formula, bg_color):
    """Merged phase-label row spanning A-N."""
    ws.row_dimensions[row].height = 22
    ws.merge_cells(f"A{row}:N{row}")
    cell = ws[f"A{row}"]
    cell.value = formula
    cell.font = make_font(bold=True, color=WHITE, size=11)
    cell.fill = make_fill(bg_color)
    cell.alignment = make_align("left")


def _write_data_row(
    ws,
    mbdd_row,
    resource_row,
    pos,
    phase_bg,
    phase_max_cell,
    is_first,
    first_date_formula,
):
    """Write one MBDD data row referencing resource tabs by row offset."""
    ws.row_dimensions[mbdd_row].height = 15

    # A: Week Of date
    if is_first and first_date_formula:
        date_inner = first_date_formula.lstrip("=")
    else:
        date_inner = f"A{mbdd_row - 1}+7"
    a_cell = ws.cell(
        row=mbdd_row,
        column=1,
        value=f'=IF({pos}<=\'Audit Setup\'!{phase_max_cell},{date_inner},"")',
    )
    a_cell.number_format = "MM/DD/YY"
    a_cell.font = make_font(size=9)
    a_cell.fill = make_fill(phase_bg)
    a_cell.alignment = make_align("center")
    a_cell.border = make_border()

    # B: Phase label (read from Resource PM as canonical reference)
    b_cell = ws.cell(
        row=mbdd_row,
        column=2,
        value=f"=IF({pos}<='Audit Setup'!{phase_max_cell},'Resource PM'!A{resource_row},\"\")",
    )
    b_cell.font = make_font(bold=True, size=9)
    b_cell.fill = make_fill(phase_bg)
    b_cell.alignment = make_align("center")
    b_cell.border = make_border()

    # C-K: Resource hours (each pulls from the corresponding tab's H column)
    for i, tab_name in enumerate(RES_TABS):
        col_idx = 3 + i
        formula = (
            f"=IF({pos}<='Audit Setup'!{phase_max_cell},"
            f"'{tab_name}'!H{resource_row},0)"
        )
        cell = ws.cell(row=mbdd_row, column=col_idx, value=formula)
        cell.font = make_font(size=9)
        cell.fill = make_fill(phase_bg)
        cell.number_format = '#,##0;-#,##0;"-"'
        cell.alignment = make_align("center")
        cell.border = make_border()

    # L: Weekly total
    l_cell = ws.cell(row=mbdd_row, column=12, value=f"=SUM(C{mbdd_row}:K{mbdd_row})")
    l_cell.font = make_font(bold=True, size=9)
    l_cell.fill = make_fill(LIGHT_GRN)
    l_cell.number_format = "#,##0"
    l_cell.alignment = make_align("center")
    l_cell.border = make_border()

    # M: Total to date (cumulative)
    if mbdd_row == MBDD_PLAN_S:
        m_formula = f"=L{mbdd_row}"
    else:
        m_formula = f"=M{mbdd_row - 1}+L{mbdd_row}"
    m_cell = ws.cell(row=mbdd_row, column=13, value=m_formula)
    m_cell.font = make_font(size=9)
    m_cell.fill = make_fill(LIGHT_GRN)
    m_cell.number_format = "#,##0"
    m_cell.alignment = make_align("center")
    m_cell.border = make_border()

    # N: Milestone/warning text
    n_cell = ws.cell(
        row=mbdd_row,
        column=14,
        value=_milestone_warning_formula(f"A{mbdd_row}"),
    )
    n_cell.font = make_font(bold=True, color=MILESTONE_FG, size=9)
    n_cell.fill = make_fill(phase_bg)
    n_cell.alignment = make_align("left", wrap=True)
    n_cell.border = make_border()


def _write_subtotal_row(ws, row, label, data_start, data_end, bg_color):
    """Phase subtotal row spanning columns A-N."""
    ws.row_dimensions[row].height = 20
    ws.merge_cells(f"A{row}:B{row}")
    label_cell = ws[f"A{row}"]
    label_cell.value = label
    label_cell.font = make_font(bold=True, color=WHITE, size=10)
    label_cell.fill = make_fill(bg_color)
    label_cell.alignment = make_align("right")

    for col_idx in range(3, 15):
        col_letter = column_letter(col_idx)
        if col_idx <= 11:
            cell = ws.cell(
                row=row,
                column=col_idx,
                value=f"=SUM({col_letter}{data_start}:{col_letter}{data_end})",
            )
        elif col_idx == 12:
            cell = ws.cell(
                row=row,
                column=col_idx,
                value=f"=SUM(L{data_start}:L{data_end})",
            )
        else:
            cell = ws.cell(row=row, column=col_idx, value="")
        cell.font = make_font(bold=True, color=WHITE, size=10)
        cell.fill = make_fill(bg_color)
        cell.number_format = "#,##0"
        cell.alignment = make_align("center")
        cell.border = make_border()


def _write_grand_total_row(ws):
    """Grand total row summing all three phase subtotals."""
    ws.row_dimensions[MBDD_GRAND].height = 22
    ws.merge_cells(f"A{MBDD_GRAND}:B{MBDD_GRAND}")
    label = ws[f"A{MBDD_GRAND}"]
    label.value = "GRAND TOTAL - ALL PHASES"
    label.font = make_font(bold=True, color=WHITE, size=11)
    label.fill = make_fill(DARK_BLUE)
    label.alignment = make_align("right")

    for col_idx in range(3, 15):
        col_letter = column_letter(col_idx)
        if col_idx <= 12:
            value = (
                f"={col_letter}{MBDD_PLAN_TOT}+{col_letter}{MBDD_FIELD_TOT}"
                f"+{col_letter}{MBDD_REP_TOT}"
            )
        else:
            value = ""
        cell = ws.cell(row=MBDD_GRAND, column=col_idx, value=value)
        cell.font = make_font(bold=True, color=WHITE, size=11)
        cell.fill = make_fill(DARK_BLUE)
        cell.number_format = "#,##0"
        cell.alignment = make_align("center")
        cell.border = make_border()


def build_mbdd_tab(wb, config, closed_range, skeleton_range):
    """
    Build the Master Budget by Date tab.

    Args:
        wb: openpyxl Workbook (must already have Audit Setup and 9 Resource tabs)
        closed_range: full closure range (accepted for interface consistency, unused here)
        skeleton_range: skeleton range (accepted for interface consistency, unused here)

    Returns:
        dict with tab metadata.
    """
    ws = wb.create_sheet("Master Budget by Date")
    ws.sheet_properties.tabColor = "FF6600"

    for col_letter, width in MBDD_COLS.items():
        ws.column_dimensions[col_letter].width = width

    _write_title_block(ws)
    _write_column_headers(ws)

    # Phase dividers with dynamic text
    _write_phase_divider(
        ws,
        MBDD_PLAN_DIV,
        f'=" \u25b8 PLANNING - Weeks 1-"&\'Audit Setup\'!$B${AS_PLAN}',
        PLAN_HDR,
    )
    _write_phase_divider(
        ws,
        MBDD_FIELD_DIV,
        (
            f'=" \u25b8 FIELDWORK - Weeks "&(\'Audit Setup\'!$B${AS_PLAN}+1)'
            f'&"-"&(\'Audit Setup\'!$B${AS_PLAN}+\'Audit Setup\'!$B${AS_FIELD})'
        ),
        FIELD_HDR,
    )
    _write_phase_divider(
        ws,
        MBDD_REP_DIV,
        (
            f'=" \u25b8 REPORTING - Weeks "&(\'Audit Setup\'!$B${AS_PLAN}'
            f'+\'Audit Setup\'!$B${AS_FIELD}+1)&"-"&\'Audit Setup\'!$B${AS_TOT}'
        ),
        REP_HDR,
    )

    # Planning rows: MBDD row r maps to resource row r+2
    for i, mbdd_row in enumerate(range(MBDD_PLAN_S, MBDD_PLAN_E + 1)):
        resource_row = mbdd_row + 2
        is_first = (i == 0)
        first_formula = f"='Audit Setup'!$C${AS_PLAN}" if is_first else None
        _write_data_row(
            ws, mbdd_row, resource_row, i + 1, PLAN_BG,
            f"$B${AS_PLAN}", is_first, first_formula,
        )

    # Fieldwork rows: MBDD row r maps to resource row r+1
    for i, mbdd_row in enumerate(range(MBDD_FIELD_S, MBDD_FIELD_E + 1)):
        resource_row = mbdd_row + 1
        is_first = (i == 0)
        first_formula = f"='Audit Setup'!$C${AS_FIELD}" if is_first else None
        _write_data_row(
            ws, mbdd_row, resource_row, i + 1, FIELD_BG,
            f"$B${AS_FIELD}", is_first, first_formula,
        )

    # Reporting rows: MBDD row r maps to resource row r (same)
    for i, mbdd_row in enumerate(range(MBDD_REP_S, MBDD_REP_E + 1)):
        resource_row = mbdd_row
        is_first = (i == 0)
        first_formula = f"='Audit Setup'!$C${AS_REP}" if is_first else None
        _write_data_row(
            ws, mbdd_row, resource_row, i + 1, REP_BG,
            f"$B${AS_REP}", is_first, first_formula,
        )

    # Subtotals and grand total
    _write_subtotal_row(
        ws, MBDD_PLAN_TOT, "PLANNING TOTAL",
        MBDD_PLAN_S, MBDD_PLAN_E, PLAN_HDR,
    )
    _write_subtotal_row(
        ws, MBDD_FIELD_TOT, "FIELDWORK TOTAL",
        MBDD_FIELD_S, MBDD_FIELD_E, FIELD_HDR,
    )
    _write_subtotal_row(
        ws, MBDD_REP_TOT, "REPORTING TOTAL",
        MBDD_REP_S, MBDD_REP_E, REP_HDR,
    )
    _write_grand_total_row(ws)

    # Conditional formatting for milestone and warning rows
    all_data_range = f"A{MBDD_PLAN_S}:N{MBDD_REP_E}"
    ws.conditional_formatting.add(
        all_data_range,
        FormulaRule(
            formula=[
                f'=AND(LEN($N{MBDD_PLAN_S})>0,'
                f'ISNUMBER(SEARCH("\u2691",$N{MBDD_PLAN_S})))'
            ],
            fill=make_fill(MS_BG),
            font=Font(name="Arial", bold=True, color=WHITE, size=9),
        ),
    )
    ws.conditional_formatting.add(
        all_data_range,
        FormulaRule(
            formula=[
                f'=AND(LEN($N{MBDD_PLAN_S})>0,'
                f'ISNUMBER(SEARCH("2 WEEKS",$N{MBDD_PLAN_S})))'
            ],
            fill=make_fill(WARN2_BG),
            font=Font(name="Arial", bold=True, color=WHITE, size=9),
        ),
    )
    ws.conditional_formatting.add(
        all_data_range,
        FormulaRule(
            formula=[
                f'=AND(LEN($N{MBDD_PLAN_S})>0,'
                f'ISNUMBER(SEARCH("NEXT WEEK",$N{MBDD_PLAN_S})))'
            ],
            fill=make_fill(WARN1_BG),
            font=Font(name="Arial", bold=True, color=BLACK, size=9),
        ),
    )

    # Pre-hide rows beyond defaults
    for i, row in enumerate(range(MBDD_PLAN_S, MBDD_PLAN_E + 1)):
        if i >= config.planning_weeks:
            ws.row_dimensions[row].hidden = True
    for i, row in enumerate(range(MBDD_FIELD_S, MBDD_FIELD_E + 1)):
        if i >= config.fieldwork_weeks:
            ws.row_dimensions[row].hidden = True
    for i, row in enumerate(range(MBDD_REP_S, MBDD_REP_E + 1)):
        if i >= config.reporting_weeks:
            ws.row_dimensions[row].hidden = True

    ws.freeze_panes = "A5"

    return {"tab_name": "Master Budget by Date"}
