"""
Builder for the nine Resource tabs.

Each resource (PM, Asst PM, Auditors 1-4, AM, QC, RE) gets its own tab
with identical structure: weekly rows across three phases (Planning,
Fieldwork, Reporting), each with hours tracking, deductions, and
milestone flags. The tab-specific elements are the title, tab color,
and display name.

All nine tabs are built by the same function, called nine times by
the orchestrator.
"""

from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation

from .config import AuditConfig

from .constants import (
    AS_FIELD,
    AS_HOL,
    AS_PLAN,
    AS_REP,
    AS_TOT,
    BLACK,
    BLUE_IN,
    BROWN,
    DARK_BLUE,
    DARK_GRN,
    DARK_RED,
    DEFAULT_FIELD,
    DEFAULT_PLAN,
    DEFAULT_REP,
    FIELD_BG,
    FIELD_DIV,
    FIELD_E,
    FIELD_HDR,
    FIELD_S,
    GRAY_LT,
    LIGHT_BLUE,
    LIGHT_GRN,
    LIGHT_RED,
    MILESTONE_BG,
    MILESTONE_FG,
    MS_ROWS,
    PLAN_BG,
    PLAN_DIV,
    PLAN_E,
    PLAN_HDR,
    PLAN_S,
    PURPLE,
    RED_F,
    REP_BG,
    REP_DIV,
    REP_E,
    REP_HDR,
    REP_S,
    RESOURCES,
    TOT_R,
    VAR_R,
    WHITE,
    YELLOW,
    column_letter,
    make_align,
    make_border,
    make_fill,
    make_font,
)


RES_COLS = {
    "A": 11, "B": 24, "C": 14, "D": 16, "E": 16,
    "F": 16, "G": 16, "H": 16, "I": 4, "J": 4, "K": 16, "L": 36,
}


def _milestone_formula(date_cell_ref):
    """
    Build the column L formula that checks 8 major milestones against
    the week's Monday date and concatenates any matches.
    """
    checks = [
        (MS_ROWS["kickoff"], "\u2691 KICKOFF"),
        (MS_ROWS["end_plan"], "\u2691 END OF PLANNING"),
        (MS_ROWS["outline"], "\u2691 OUTLINE TO EXEC"),
        (MS_ROWS["ontarget"], "\u2691 ON-TARGET"),
        (MS_ROWS["draft"], "\u2691 DRAFT TO EXEC"),
        (MS_ROWS["exit"], "\u2691 EXIT CONFERENCE"),
        (MS_ROWS["mgmt"], "\u2691 MGMT RESPONSES DUE"),
        (MS_ROWS["release"], "\u2691 REPORT RELEASE"),
    ]
    parts = [
        f'IF(AND(\'Audit Setup\'!$C${r}>={date_cell_ref},'
        f'\'Audit Setup\'!$C${r}<={date_cell_ref}+4),"{label} ","")'
        for r, label in checks
    ]
    inner = ",".join(parts)
    return (
        f'=IF(NOT(ISNUMBER({date_cell_ref})),"",'
        f'TRIM(CONCATENATE({inner})))'
    )


def _write_title_block(ws, display_name):
    """Title row and two input rows for resource name and project name."""
    ws.merge_cells("A1:L1")
    title = ws["A1"]
    title.value = f"Audit Resource Tracker | v4 | {display_name}"
    title.font = make_font(bold=True, color=WHITE, size=12)
    title.fill = make_fill(DARK_BLUE)
    title.alignment = make_align("center")
    ws.row_dimensions[1].height = 24

    for row, label in [(2, "Resource Name:"), (3, "Audit / Project:")]:
        ws.row_dimensions[row].height = 18
        ws.merge_cells(f"A{row}:B{row}")
        label_cell = ws[f"A{row}"]
        label_cell.value = label
        label_cell.font = make_font(bold=True, size=10)
        label_cell.alignment = make_align()

        ws.merge_cells(f"C{row}:L{row}")
        value_cell = ws[f"C{row}"]
        value_cell.value = display_name if row == 2 else "Enter audit name"
        value_cell.font = make_font(color=BLUE_IN, size=10)
        value_cell.fill = make_fill(LIGHT_BLUE)
        value_cell.alignment = make_align()


def _write_note_bar(ws):
    """Instruction bar spanning columns A-L."""
    ws.merge_cells("A4:L4")
    note = ws["A4"]
    note.value = "Phase structure driven by 'Audit Setup'. Blue = manual entry."
    note.font = make_font(italic=True, size=8, color="595959")
    note.fill = make_fill(GRAY_LT)
    note.alignment = make_align("left", wrap=True)
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 6


def _write_column_headers(ws):
    """Row 6 — 12 column headers."""
    ws.row_dimensions[6].height = 42
    headers = [
        (1, "Week\n#", DARK_BLUE, WHITE),
        (2, "Task\nAssignment", PLAN_HDR, WHITE),
        (3, "Week Of\n(Mon)", DARK_BLUE, WHITE),
        (4, "Hours\nAvail.", PLAN_HDR, WHITE),
        (5, "(-) Business\nHolidays", DARK_RED, WHITE),
        (6, "(-) Skeleton\nDays", BROWN, WHITE),
        (7, "(-) Leave\nPlanned", DARK_RED, WHITE),
        (8, "Total\nAvail.", DARK_GRN, WHITE),
        (9, "", WHITE, WHITE),
        (10, "", WHITE, WHITE),
        (11, "Actual Hrs", DARK_BLUE, WHITE),
        (12, "Milestones\nThis Week", PURPLE, WHITE),
    ]
    for col_idx, label, bg, fg in headers:
        cell = ws.cell(row=6, column=col_idx, value=label)
        cell.font = make_font(bold=True, color=fg, size=9)
        cell.fill = make_fill(bg)
        cell.alignment = make_align("center", wrap=True)
        if col_idx not in (9, 10):
            cell.border = make_border()


def _write_phase_divider(ws, row, formula, bg_color):
    """Merged phase-label row spanning A-L."""
    ws.row_dimensions[row].height = 22
    ws.merge_cells(f"A{row}:L{row}")
    cell = ws[f"A{row}"]
    cell.value = formula
    cell.font = make_font(bold=True, color=WHITE, size=11)
    cell.fill = make_fill(bg_color)
    cell.alignment = make_align("left")


def _write_data_row(
    ws,
    row,
    pos,
    phase_bg,
    max_weeks_cell,
    prefix,
    first_date_formula,
    closed_range,
    skeleton_range,
):
    """Write a single data row. pos is 1-indexed within its phase."""
    ws.row_dimensions[row].height = 15

    # A: Phase label
    a_cell = ws.cell(
        row=row,
        column=1,
        value=f'=IF({pos}<=\'Audit Setup\'!{max_weeks_cell},"{prefix}-"&{pos},"")',
    )
    a_cell.font = make_font(bold=True, size=9)
    a_cell.fill = make_fill(phase_bg)
    a_cell.alignment = make_align("center")
    a_cell.border = make_border()

    # B: Task assignment (manual)
    b_cell = ws.cell(row=row, column=2, value="")
    b_cell.font = make_font(size=9, italic=True, color="595959")
    b_cell.fill = make_fill(LIGHT_BLUE)
    b_cell.alignment = make_align("left")
    b_cell.border = make_border()

    # C: Week Of (Monday)
    if pos == 1:
        c_value = first_date_formula
    else:
        c_value = f"=C{row - 1}+7"
    c_cell = ws.cell(row=row, column=3, value=c_value)
    c_cell.number_format = "MM/DD/YY"
    c_cell.font = make_font(size=9)
    c_cell.fill = make_fill(phase_bg)
    c_cell.alignment = make_align("center")
    c_cell.border = make_border()

    # D: Hours available (manual)
    d_cell = ws.cell(row=row, column=4, value=0)
    d_cell.font = make_font(color=BLUE_IN, size=9)
    d_cell.fill = make_fill(LIGHT_BLUE)
    d_cell.number_format = '#,##0.0;[Red]-#,##0.0;"-"'
    d_cell.alignment = make_align("center")
    d_cell.border = make_border()

    # E: Business holidays deduction
    e_formula = (
        f'=-COUNTIFS({closed_range},">="&C{row},'
        f'{closed_range},"<="&(C{row}+4))*\'Audit Setup\'!$B${AS_HOL}'
    )
    e_cell = ws.cell(row=row, column=5, value=e_formula)
    e_cell.font = make_font(color=RED_F, size=9)
    e_cell.fill = make_fill(phase_bg)
    e_cell.number_format = '#,##0.0;-#,##0.0;"-"'
    e_cell.alignment = make_align("center")
    e_cell.border = make_border()

    # F: Skeleton days deduction
    f_formula = (
        f'=-COUNTIFS({skeleton_range},">="&C{row},'
        f'{skeleton_range},"<="&(C{row}+4))*\'Audit Setup\'!$B${AS_HOL}'
    )
    f_cell = ws.cell(row=row, column=6, value=f_formula)
    f_cell.font = make_font(color=BROWN, size=9)
    f_cell.fill = make_fill(phase_bg)
    f_cell.number_format = '#,##0.0;-#,##0.0;"-"'
    f_cell.alignment = make_align("center")
    f_cell.border = make_border()

    # G: Leave (manual)
    g_cell = ws.cell(row=row, column=7, value=0)
    g_cell.font = make_font(color=RED_F, size=9)
    g_cell.fill = make_fill(LIGHT_BLUE)
    g_cell.number_format = '#,##0.0;[Red]-#,##0.0;"-"'
    g_cell.alignment = make_align("center")
    g_cell.border = make_border()

    # H: Total available
    h_cell = ws.cell(
        row=row, column=8, value=f"=CEILING(MAX(0,SUM(D{row}:G{row})),1)"
    )
    h_cell.font = make_font(bold=True, size=9)
    h_cell.fill = make_fill(LIGHT_GRN)
    h_cell.number_format = "#,##0"
    h_cell.alignment = make_align("center")
    h_cell.border = make_border()

    # I, J: Spacers
    for col_idx in (9, 10):
        ws.cell(row=row, column=col_idx).fill = make_fill(GRAY_LT)

    # K: Actual hours (manual)
    k_cell = ws.cell(row=row, column=11, value=0)
    k_cell.font = make_font(color=BLUE_IN, size=9)
    k_cell.fill = make_fill(LIGHT_BLUE)
    k_cell.number_format = '#,##0.0;-#,##0.0;"-"'
    k_cell.alignment = make_align("center")
    k_cell.border = make_border()

    # L: Milestone flags
    l_cell = ws.cell(row=row, column=12, value=_milestone_formula(f"C{row}"))
    l_cell.font = make_font(bold=True, color=MILESTONE_FG, size=9)
    l_cell.fill = make_fill(phase_bg)
    l_cell.alignment = make_align("left", wrap=True)
    l_cell.border = make_border()


def _write_totals_row(ws):
    """Row 95 — sums across all data rows for each numeric column."""
    ws.row_dimensions[TOT_R].height = 22
    ws.merge_cells(f"A{TOT_R}:C{TOT_R}")
    label = ws.cell(row=TOT_R, column=1, value="TOTALS")
    label.font = make_font(bold=True, color=WHITE, size=11)
    label.fill = make_fill(DARK_BLUE)
    label.alignment = make_align("center")
    label.border = make_border()

    def sum_range(col):
        return (
            f"SUM({col}{PLAN_S}:{col}{PLAN_E},"
            f"{col}{FIELD_S}:{col}{FIELD_E},"
            f"{col}{REP_S}:{col}{REP_E})"
        )

    columns_to_total = [
        (4, BLUE_IN, LIGHT_BLUE),
        (5, RED_F, LIGHT_RED),
        (6, BROWN, YELLOW),
        (7, RED_F, LIGHT_RED),
        (8, DARK_GRN, LIGHT_GRN),
        (11, BLUE_IN, LIGHT_BLUE),
    ]
    for col_idx, font_color, fill_color in columns_to_total:
        col_letter = column_letter(col_idx)
        cell = ws.cell(row=TOT_R, column=col_idx, value=f"={sum_range(col_letter)}")
        cell.font = make_font(bold=True, color=font_color, size=11)
        cell.fill = make_fill(fill_color)
        cell.number_format = "#,##0"
        cell.alignment = make_align("center")
        cell.border = make_border()

    for col_idx in (9, 10):
        ws.cell(row=TOT_R, column=col_idx).fill = make_fill(GRAY_LT)
    ws.cell(row=TOT_R, column=12).fill = make_fill(GRAY_LT)


def _write_variance_row(ws):
    """Row 96 — variance label and formula."""
    ws.row_dimensions[VAR_R].height = 18
    ws.merge_cells(f"A{VAR_R}:G{VAR_R}")
    label = ws[f"A{VAR_R}"]
    label.value = "Variance: Actual (K) minus Total Available (H)"
    label.font = make_font(bold=True, italic=True, size=9)
    label.fill = make_fill(GRAY_LT)
    label.alignment = make_align("right")

    variance = ws.cell(row=VAR_R, column=8, value=f"=K{TOT_R}-H{TOT_R}")
    variance.font = make_font(bold=True, size=11)
    variance.number_format = '#,##0;[Red]-#,##0;"-"'
    variance.alignment = make_align("center")
    variance.border = make_border()
    variance.fill = make_fill(GRAY_LT)

    ws.merge_cells(f"I{VAR_R}:L{VAR_R}")
    note = ws[f"I{VAR_R}"]
    note.value = "Positive = over budget | Negative = under budget"
    note.font = make_font(italic=True, size=8, color="595959")
    note.fill = make_fill(GRAY_LT)
    note.alignment = make_align("left")


def build_resource_tab(wb, config, tab_name, display_name, tab_color, closed_range, skeleton_range):
    """
    Build one Resource tab in the given workbook.

    Args:
        wb: openpyxl Workbook
        tab_name: internal tab name (e.g., "Resource PM")
        display_name: label shown in titles (e.g., "PM")
        tab_color: hex color string for the tab color
        closed_range: cell range reference for full closures
        skeleton_range: cell range reference for skeleton days

    Returns:
        The created worksheet object.
    """
    ws = wb.create_sheet(tab_name)
    ws.sheet_properties.tabColor = tab_color

    for col_letter, width in RES_COLS.items():
        ws.column_dimensions[col_letter].width = width

    _write_title_block(ws, display_name)
    _write_note_bar(ws)
    _write_column_headers(ws)

    # Phase dividers with dynamic labels
    _write_phase_divider(
        ws,
        PLAN_DIV,
        f'=" PLANNING - Weeks 1-"&\'Audit Setup\'!$B${AS_PLAN}',
        PLAN_HDR,
    )
    _write_phase_divider(
        ws,
        FIELD_DIV,
        (
            f'=" FIELDWORK - Weeks "&(\'Audit Setup\'!$B${AS_PLAN}+1)'
            f'&"-"&(\'Audit Setup\'!$B${AS_PLAN}+\'Audit Setup\'!$B${AS_FIELD})'
        ),
        FIELD_HDR,
    )
    _write_phase_divider(
        ws,
        REP_DIV,
        (
            f'=" REPORTING - Weeks "&(\'Audit Setup\'!$B${AS_PLAN}'
            f'+\'Audit Setup\'!$B${AS_FIELD}+1)&"-"&\'Audit Setup\'!$B${AS_TOT}'
        ),
        REP_HDR,
    )

    # Planning rows
    for i, row in enumerate(range(PLAN_S, PLAN_E + 1)):
        _write_data_row(
            ws, row, i + 1, PLAN_BG, f"$B${AS_PLAN}", "PL",
            f"='Audit Setup'!$C${AS_PLAN}",
            closed_range, skeleton_range,
        )

    # Fieldwork rows
    for i, row in enumerate(range(FIELD_S, FIELD_E + 1)):
        _write_data_row(
            ws, row, i + 1, FIELD_BG, f"$B${AS_FIELD}", "FW",
            f"='Audit Setup'!$C${AS_FIELD}",
            closed_range, skeleton_range,
        )

    # Reporting rows
    for i, row in enumerate(range(REP_S, REP_E + 1)):
        _write_data_row(
            ws, row, i + 1, REP_BG, f"$B${AS_REP}", "RP",
            f"='Audit Setup'!$C${AS_REP}",
            closed_range, skeleton_range,
        )

    # Conditional formatting: highlight milestone rows
    all_data_range = f"A{PLAN_S}:L{REP_E}"
    ws.conditional_formatting.add(
        all_data_range,
        FormulaRule(
            formula=[f'=AND(LEN($L{PLAN_S})>0,ISNUMBER(SEARCH("\u2691",$L{PLAN_S})))'],
            fill=make_fill(MILESTONE_BG),
            font=Font(name="Arial", bold=True, color=MILESTONE_FG, size=9),
        ),
    )

    # Data validation on column D
    dv = DataValidation(
        type="decimal",
        operator="between",
        formula1="0",
        formula2="40",
        error="0-40 hrs only",
        errorTitle="Invalid Entry",
        showErrorMessage=True,
    )
    ws.add_data_validation(dv)
    dv.sqref = f"D{PLAN_S}:D{REP_E}"

    _write_totals_row(ws)
    _write_variance_row(ws)

    ws.freeze_panes = "A7"

    # Pre-hide rows beyond defaults
    for i, row in enumerate(range(PLAN_S, PLAN_E + 1)):
        if i >= config.planning_weeks:
            ws.row_dimensions[row].hidden = True
    for i, row in enumerate(range(FIELD_S, FIELD_E + 1)):
        if i >= config.fieldwork_weeks:
            ws.row_dimensions[row].hidden = True
    for i, row in enumerate(range(REP_S, REP_E + 1)):
        if i >= config.reporting_weeks:
            ws.row_dimensions[row].hidden = True

    return ws


def build_all_resource_tabs(wb, config, closed_range, skeleton_range):
    """
    Build all 9 resource tabs using the RESOURCES list from constants.
    Returns a dict mapping display names to worksheet objects.
    """
    result = {}
    for tab_name, display_name, tab_color in RESOURCES:
        ws = build_resource_tab(
            wb, config, tab_name, display_name, tab_color,
            closed_range, skeleton_range,
        )
        result[display_name] = ws
    return result
