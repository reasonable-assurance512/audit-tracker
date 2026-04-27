"""
Builder for the Audit Setup tab.

Produces the central configuration tab: anchor dates (kickoff, release),
phase configuration (planning/fieldwork/reporting weeks), and the
milestone calendar. Every downstream tab reads from this tab's cells,
so its row layout is fixed and documented in constants.py.
"""

from datetime import date

from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font

from .config import AuditConfig

from .constants import (
    AS_FIELD,
    AS_KICK,
    AS_PLAN,
    AS_REL,
    AS_REP,
    AS_TOT,
    BLACK,
    BLUE_IN,
    FIELD_BG,
    FIELD_HDR,
    GRAY_LT,
    GRAY_MD,
    LIGHT_BLUE,
    LIGHT_GRN,
    MS_ROWS,
    PLAN_BG,
    PLAN_HDR,
    PURPLE,
    REP_BG,
    REP_HDR,
    WARN_BG,
    WARN_RED,
    WHITE,
    column_letter,
    make_align,
    make_border,
    make_fill,
    make_font,
)


def _section_header(ws, row, text, bg_color, span=5):
    """Write a full-width section header bar."""
    ws.merge_cells(f"A{row}:{column_letter(span)}{row}")
    cell = ws[f"A{row}"]
    cell.value = text
    cell.font = make_font(bold=True, color=WHITE, size=10)
    cell.fill = make_fill(bg_color)
    cell.alignment = make_align()
    ws.row_dimensions[row].height = 18


def _column_headers(ws, row, items, height=18):
    """Write a row of column headers with consistent styling."""
    ws.row_dimensions[row].height = height
    for col_idx, (label, bg, fg) in enumerate(items, 1):
        cell = ws.cell(row=row, column=col_idx, value=label)
        cell.font = make_font(bold=True, color=fg, size=9)
        cell.fill = make_fill(bg)
        cell.alignment = make_align("center")
        cell.border = make_border()


def _input_row(ws, row, label, value, number_format=None, note="", note_bold=False):
    """Write a labeled input row with value cell and optional note span."""
    ws.row_dimensions[row].height = 18
    label_cell = ws.cell(row=row, column=1, value=label)
    label_cell.font = make_font(bold=True, size=10)
    label_cell.border = make_border()
    label_cell.alignment = make_align()

    value_cell = ws.cell(row=row, column=2, value=value)
    value_cell.font = make_font(bold=True, color=BLUE_IN, size=11)
    value_cell.fill = make_fill(LIGHT_BLUE)
    value_cell.border = make_border()
    value_cell.alignment = make_align("center")
    if number_format:
        value_cell.number_format = number_format

    if note:
        ws.merge_cells(f"C{row}:E{row}")
        note_cell = ws.cell(row=row, column=3, value=note)
        note_cell.font = make_font(
            bold=note_bold,
            size=9,
            color="595959" if not note_bold else BLACK,
        )
        note_cell.alignment = make_align("left", wrap=True)
        note_cell.border = make_border()


def build_setup_tab(wb, config, closed_range, skeleton_range):
    """
    Build the Audit Setup tab in the given workbook.

    Args:
        wb: openpyxl Workbook instance (must have at least one sheet already)
        closed_range: cell range string for full closure dates
        skeleton_range: cell range string for skeleton crew dates (unused here
            but accepted for interface consistency across tab builders)

    Returns:
        dict with tab metadata for downstream tabs
    """
    ws = wb.create_sheet("Audit Setup")
    ws.sheet_properties.tabColor = PURPLE

    for col_letter, width in [("A", 36), ("B", 16), ("C", 16), ("D", 16), ("E", 46)]:
        ws.column_dimensions[col_letter].width = width

    # ── Title row ───────────────────────────────────────────
    ws.merge_cells("A1:E1")
    title = ws["A1"]
    title.value = "Audit Configuration & Milestone Calculator | v4"
    title.font = make_font(bold=True, color=WHITE, size=13)
    title.fill = make_fill(PURPLE)
    title.alignment = make_align("center")
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 6

    # ── Section 1: Anchor Dates ────────────────────────────
    _section_header(ws, 3, "ANCHOR DATES", PURPLE)
    _column_headers(
        ws,
        4,
        [
            ("Field", GRAY_MD, WHITE),
            ("Value", GRAY_MD, WHITE),
            ("Day of Week / Calc", GRAY_MD, WHITE),
            ("", GRAY_MD, WHITE),
            ("Validation / Notes", GRAY_MD, WHITE),
        ],
    )

    # Row 5: Kickoff date
    ws.row_dimensions[5].height = 18
    label = ws.cell(row=5, column=1, value="Kickoff / Project Launch Date")
    label.font = make_font(bold=True, size=10)
    label.border = make_border()
    label.alignment = make_align()

    kickoff_cell = ws.cell(row=5, column=2, value=config.kickoff_date)
    kickoff_cell.font = make_font(bold=True, color=BLUE_IN, size=11)
    kickoff_cell.fill = make_fill(LIGHT_BLUE)
    kickoff_cell.number_format = "MM/DD/YYYY"
    kickoff_cell.border = make_border()
    kickoff_cell.alignment = make_align("center")

    ws.merge_cells("C5:D5")
    dow = ws.cell(row=5, column=3, value='=TEXT(B5,"dddd")')
    dow.font = make_font(size=10)
    dow.alignment = make_align("center")
    dow.border = make_border()

    validation = ws.cell(
        row=5,
        column=5,
        value='=IF(WEEKDAY(B5,2)=5,"\u26a0 Kickoff is a Friday","\u2713 OK")',
    )
    validation.alignment = make_align()
    validation.border = make_border()

    # Row 6: Report Release date (formula-driven)
    ws.row_dimensions[6].height = 18
    label = ws.cell(row=6, column=1, value="Report Release Date (auto-calculated)")
    label.font = make_font(bold=True, size=10)
    label.border = make_border()
    label.alignment = make_align()

    release_cell = ws.cell(
        row=6,
        column=2,
        value=f"=WORKDAY(B5-WEEKDAY(B5,2)+1+(B{AS_TOT}*7)-3,1,{closed_range})",
    )
    release_cell.font = make_font(bold=True, size=11)
    release_cell.fill = make_fill(LIGHT_GRN)
    release_cell.number_format = "MM/DD/YYYY"
    release_cell.border = make_border()
    release_cell.alignment = make_align("center")

    ws.merge_cells("C6:D6")
    dow = ws.cell(row=6, column=3, value='=TEXT(B6,"dddd")')
    dow.font = make_font(size=10)
    dow.alignment = make_align("center")
    dow.border = make_border()

    validation = ws.cell(
        row=6,
        column=5,
        value='=IF(WEEKDAY(B6,2)<>1,"\u26a0 Release is not Monday","\u2713 OK")',
    )
    validation.border = make_border()
    validation.alignment = make_align()

    # Row 7: Hours per holiday (manual input)
    _input_row(
        ws,
        7,
        "Standard Hours Lost per Holiday/Skeleton Day",
        config.hours_per_holiday,
        "0",
        "Deducted from resource tabs per closure. Default: 8.",
    )

    # Row 8: On-Target buffer
    ws.row_dimensions[8].height = 18
    label = ws.cell(
        row=8, column=1, value="Weeks Before End of Fieldwork for On-Target Meeting"
    )
    label.font = make_font(bold=True, size=10)
    label.border = make_border()
    label.alignment = make_align()

    buffer_cell = ws.cell(row=8, column=2, value=config.on_target_buffer)
    buffer_cell.font = make_font(bold=True, color=BLUE_IN, size=11)
    buffer_cell.fill = make_fill(LIGHT_BLUE)
    buffer_cell.number_format = "0"
    buffer_cell.border = make_border()
    buffer_cell.alignment = make_align("center")

    ws.merge_cells("C8:E8")
    note_cell = ws.cell(
        row=8,
        column=3,
        value='="On-Target = "&TEXT(D13-4-$B$8*7,"MM/DD/YY dddd")&" | Buffer: "&$B$8&" wk(s)"',
    )
    note_cell.alignment = make_align("left")
    note_cell.border = make_border()

    ws.row_dimensions[9].height = 8

    # Conditional formatting: release date not Monday warning
    ws.conditional_formatting.add(
        "E6",
        FormulaRule(
            formula=["=WEEKDAY($B$6,2)<>1"],
            fill=make_fill(WARN_BG),
            font=Font(name="Arial", color=WARN_RED, bold=True, size=9),
        ),
    )

    # ── Section 2: Phase Configuration ──────────────────────
    _section_header(ws, 10, "PHASE CONFIGURATION", PLAN_HDR)
    _column_headers(
        ws,
        11,
        [
            ("Phase", GRAY_MD, WHITE),
            ("Weeks Allocated", GRAY_MD, WHITE),
            ("Start (Mon)", GRAY_MD, WHITE),
            ("End (Fri)", GRAY_MD, WHITE),
            ("Notes / Validation", GRAY_MD, WHITE),
        ],
    )

    phases = [
        (
            AS_PLAN,
            "PLANNING",
            config.planning_weeks,
            PLAN_HDR,
            PLAN_BG,
            "=B5-WEEKDAY(B5,2)+1",
            f"=C{AS_PLAN}+(B{AS_PLAN}-1)*7+4",
            "Start snaps to Monday of kickoff week",
        ),
        (
            AS_FIELD,
            "FIELDWORK",
            config.fieldwork_weeks,
            FIELD_HDR,
            FIELD_BG,
            f"=D{AS_PLAN}+3",
            f"=C{AS_FIELD}+(B{AS_FIELD}-1)*7+4",
            "",
        ),
        (
            AS_REP,
            "REPORTING",
            config.reporting_weeks,
            REP_HDR,
            REP_BG,
            f"=D{AS_FIELD}+3",
            f"=C{AS_REP}+(B{AS_REP}-1)*7+4",
            "Begins Monday after Fieldwork ends",
        ),
    ]

    for row, phase, weeks, header_color, bg_color, start_formula, end_formula, note in phases:
        ws.row_dimensions[row].height = 20
        phase_cell = ws.cell(row=row, column=1, value=phase)
        phase_cell.font = make_font(bold=True, color=WHITE, size=11)
        phase_cell.fill = make_fill(header_color)
        phase_cell.alignment = make_align()
        phase_cell.border = make_border()

        weeks_cell = ws.cell(row=row, column=2, value=weeks)
        weeks_cell.font = make_font(bold=True, color=BLUE_IN, size=12)
        weeks_cell.fill = make_fill(LIGHT_BLUE)
        weeks_cell.alignment = make_align("center")
        weeks_cell.border = make_border()
        weeks_cell.number_format = '0" wks"'

        start_cell = ws.cell(row=row, column=3, value=start_formula)
        start_cell.font = make_font(size=10)
        start_cell.number_format = "MM/DD/YY"
        start_cell.fill = make_fill(bg_color)
        start_cell.alignment = make_align("center")
        start_cell.border = make_border()

        end_cell = ws.cell(row=row, column=4, value=end_formula)
        end_cell.font = make_font(size=10)
        end_cell.number_format = "MM/DD/YY"
        end_cell.fill = make_fill(bg_color)
        end_cell.alignment = make_align("center")
        end_cell.border = make_border()

        if note:
            note_cell = ws.cell(row=row, column=5, value=note)
            note_cell.font = make_font(italic=True, size=9, color="595959")
            note_cell.alignment = make_align()
            note_cell.border = make_border()

    # Fieldwork row gets a validation formula in column E (writing window check)
    writing_window_formula = (
        f'=IF(C{MS_ROWS["draft"]}-C{MS_ROWS["ontarget"]}<28,'
        f'"\u26a0 Writing window short","\u2713 OK")'
    )
    fw_validation = ws.cell(row=AS_FIELD, column=5, value=writing_window_formula)
    fw_validation.font = make_font(bold=True, size=9)
    fw_validation.alignment = make_align(wrap=True)
    fw_validation.border = make_border()
    ws.row_dimensions[AS_FIELD].height = 28

    ws.conditional_formatting.add(
        f"E{AS_FIELD}",
        FormulaRule(
            formula=[f'=C{MS_ROWS["draft"]}-C{MS_ROWS["ontarget"]}<28'],
            fill=make_fill(WARN_BG),
            font=Font(name="Arial", color=WARN_RED, bold=True, size=9),
        ),
    )

    # Total weeks row
    ws.row_dimensions[15].height = 8
    ws.row_dimensions[AS_TOT].height = 18

    total_label = ws.cell(row=AS_TOT, column=1, value="TOTAL AUDIT WEEKS")
    total_label.font = make_font(bold=True, size=10)
    total_label.border = make_border()
    total_label.fill = make_fill(GRAY_LT)
    total_label.alignment = make_align()

    total_cell = ws.cell(row=AS_TOT, column=2, value=f"=SUM(B{AS_PLAN}:B{AS_REP})")
    total_cell.font = make_font(bold=True, size=12)
    total_cell.fill = make_fill(GRAY_MD)
    total_cell.border = make_border()
    total_cell.alignment = make_align("center")
    total_cell.number_format = '0" wks"'

    ws.merge_cells(f"C{AS_TOT}:E{AS_TOT}")
    total_note = ws.cell(
        row=AS_TOT,
        column=3,
        value=f'="Audit: "&TEXT(C{AS_PLAN},"MM/DD/YY")&" - "&TEXT(D{AS_REP},"MM/DD/YY")',
    )
    total_note.alignment = make_align()
    total_note.border = make_border()

    ws.row_dimensions[17].height = 8

    # ── Section 3: Milestone Calendar ──────────────────────
    _section_header(ws, 18, "MILESTONE CALENDAR", PURPLE)
    _column_headers(
        ws,
        19,
        [
            ("Milestone", GRAY_MD, WHITE),
            ("", GRAY_MD, WHITE),
            ("Date", GRAY_MD, WHITE),
            ("Day of Week", GRAY_MD, WHITE),
            ("Notes / Validation", GRAY_MD, WHITE),
        ],
    )

    milestones = [
        (MS_ROWS["kickoff"], "Kickoff / Project Launch", f"=B{AS_KICK}", "Anchor", True),
        (MS_ROWS["end_plan"], "End of Planning", f"=D{AS_PLAN}", "Last Friday of Planning", False),
        (
            MS_ROWS["outline"],
            "Outline Sent to Exec",
            f"=WORKDAY(C{MS_ROWS['ontarget']},-2,{closed_range})",
            "2 business days before On-Target",
            False,
        ),
        (
            MS_ROWS["ontarget"],
            "ON-TARGET MEETING",
            f"=D{AS_FIELD}-4-$B$8*7",
            "Monday of week B8 weeks before end of Fieldwork",
            True,
        ),
        (MS_ROWS["writing"], "Writing Period Begins", f"=C{MS_ROWS['ontarget']}+1", "Day after On-Target", False),
        (
            MS_ROWS["draft"],
            "Draft Sent to Exec",
            f"=WORKDAY(C{MS_ROWS['exit']},-3,{closed_range})",
            "3 business days before Exit Conference",
            False,
        ),
        (
            MS_ROWS["exit"],
            "Exit Conference",
            f"=WORKDAY(C{MS_ROWS['mgmt']},-10,{closed_range})",
            "10 business days before Mgmt Responses Due",
            False,
        ),
        (MS_ROWS["mgmt"], "Management Responses Due", f"=B{AS_REL}-7", "Monday of release week minus 1 week", False),
        (MS_ROWS["final"], "Final Draft to Exec", f"=B{AS_REL}-6", "Tuesday of release week", False),
        (MS_ROWS["exec_app"], "Exec Approves Final Draft", f"=B{AS_REL}-5", "Wednesday of release week", False),
        (MS_ROWS["pm_am"], "PM / AM Review Prototype", f"=B{AS_REL}-5", "Wednesday-Thursday of release week", False),
        (MS_ROWS["editors"], "Editors Set Up Release Emails", f"=B{AS_REL}-3", "Friday of release week", False),
        (MS_ROWS["release"], "REPORT RELEASE", f"=B{AS_REL}", "Anchor", True),
    ]

    for row, name, formula, description, is_anchor in milestones:
        ws.row_dimensions[row].height = 18
        bg = LIGHT_BLUE if is_anchor else WHITE

        name_cell = ws.cell(row=row, column=1, value=name)
        name_cell.font = make_font(bold=is_anchor, size=10 if is_anchor else 9)
        name_cell.fill = make_fill(bg)
        name_cell.alignment = make_align()
        name_cell.border = make_border()

        ws.cell(row=row, column=2).border = make_border()

        date_cell = ws.cell(row=row, column=3, value=formula)
        date_cell.font = make_font(bold=is_anchor, size=10)
        date_cell.number_format = "MM/DD/YYYY"
        date_cell.fill = make_fill(bg)
        date_cell.alignment = make_align("center")
        date_cell.border = make_border()

        dow_cell = ws.cell(
            row=row,
            column=4,
            value=f'=IF(ISNUMBER(C{row}),TEXT(C{row},"dddd"),"")',
        )
        dow_cell.font = make_font(size=9)
        dow_cell.fill = make_fill(bg)
        dow_cell.alignment = make_align("center")
        dow_cell.border = make_border()

        desc_cell = ws.cell(row=row, column=5, value=description)
        desc_cell.font = make_font(italic=True, size=9, color="595959")
        desc_cell.fill = make_fill(bg)
        desc_cell.alignment = make_align()
        desc_cell.border = make_border()

    # Draft row gets a writing-window validation
    draft_validation = ws.cell(
        row=MS_ROWS["draft"],
        column=5,
        value=(
            f'=IF(C{MS_ROWS["draft"]}-C{MS_ROWS["ontarget"]}<28,'
            f'"\u26a0 Writing window short","")'
        ),
    )
    draft_validation.border = make_border()
    ws.row_dimensions[MS_ROWS["draft"]].height = 24

    # Release row gets a not-Monday validation
    release_validation = ws.cell(
        row=MS_ROWS["release"],
        column=5,
        value=f'=IF(WEEKDAY(B{AS_REL},2)<>1,"\u26a0 Release not Monday","\u2713 OK")',
    )
    release_validation.border = make_border()

    # Conditional formatting for draft row (writing window warning)
    ws.conditional_formatting.add(
        f'A{MS_ROWS["draft"]}:E{MS_ROWS["draft"]}',
        FormulaRule(
            formula=[f'=C{MS_ROWS["draft"]}-C{MS_ROWS["ontarget"]}<28'],
            fill=make_fill(WARN_BG),
            font=Font(name="Arial", color=WARN_RED, bold=True, size=9),
        ),
    )

    # Conditional formatting for release row (not-Monday warning)
    ws.conditional_formatting.add(
        f'A{MS_ROWS["release"]}:E{MS_ROWS["release"]}',
        FormulaRule(
            formula=[f"=WEEKDAY($B${AS_REL},2)<>1"],
            fill=make_fill(WARN_BG),
            font=Font(name="Arial", color=WARN_RED, bold=True, size=9),
        ),
    )

    ws.freeze_panes = "A5"

    return {"tab_name": "Audit Setup"}
