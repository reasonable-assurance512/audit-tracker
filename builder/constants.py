"""
Shared constants and helper functions used across builder modules.

Defines color tokens, row/column layout constants, and small helper
functions for creating openpyxl style objects. Any value or helper
needed by more than one tab module belongs here.
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Color tokens ──────────────────────────────────────────
DARK_BLUE = "1F3864"
MED_BLUE = "2E5090"
LIGHT_BLUE = "DCE6F1"
DARK_RED = "C00000"
LIGHT_RED = "FFE0E0"
DARK_GRN = "375623"
LIGHT_GRN = "EBF1DE"
YELLOW = "FFF2CC"
ORANGE = "ED7D31"
BROWN = "7B3F00"
GRAY_LT = "F2F2F2"
GRAY_MD = "D9D9D9"
WHITE = "FFFFFF"
BLACK = "000000"
BLUE_IN = "0000CD"
RED_F = "C00000"
PURPLE = "7030A0"
LIGHT_PURPLE = "E8DDEF"
PLAN_HDR = "2C4770"
PLAN_BG = "D6E4F0"
FIELD_HDR = "1D5C2C"
FIELD_BG = "D9F0D9"
REP_HDR = "7B3F00"
REP_BG = "FFF0CC"
MILESTONE_BG = "FFD966"
MILESTONE_FG = "7B3F00"
WARN_BG = "FFE0E0"
WARN_RED = "C00000"
TEAL = "006B6B"
TEAL_LT = "E0F0F0"
AM_CLR = "7030A0"
QC_CLR = "C55A11"
RE_CLR = "006B6B"
MS_BG = "1F3864"
WARN1_BG = "FFC000"
WARN2_BG = "FF9900"


# ── Resource tab layout constants ─────────────────────────
MAX_PLAN = 25
MAX_FIELD = 40
MAX_REP = 20
PLAN_DIV = 7
PLAN_S = 8
PLAN_E = PLAN_S + MAX_PLAN - 1
FIELD_DIV = 33
FIELD_S = 34
FIELD_E = FIELD_S + MAX_FIELD - 1
REP_DIV = 74
REP_S = 75
REP_E = REP_S + MAX_REP - 1
TOT_R = 95
VAR_R = 96
DEFAULT_PLAN = 4
DEFAULT_FIELD = 16
DEFAULT_REP = 4

"""
Shared constants and helper functions used across builder modules.

Defines color tokens, row/column layout constants, and small helper
functions for creating openpyxl style objects. Any value or helper
needed by more than one tab module belongs here.
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Color tokens ──────────────────────────────────────────
DARK_BLUE = "1F3864"
MED_BLUE = "2E5090"
LIGHT_BLUE = "DCE6F1"
DARK_RED = "C00000"
LIGHT_RED = "FFE0E0"
DARK_GRN = "375623"
LIGHT_GRN = "EBF1DE"
YELLOW = "FFF2CC"
ORANGE = "ED7D31"
BROWN = "7B3F00"
GRAY_LT = "F2F2F2"
GRAY_MD = "D9D9D9"
WHITE = "FFFFFF"
BLACK = "000000"
BLUE_IN = "0000CD"
RED_F = "C00000"
PURPLE = "7030A0"
LIGHT_PURPLE = "E8DDEF"
PLAN_HDR = "2C4770"
PLAN_BG = "D6E4F0"
FIELD_HDR = "1D5C2C"
FIELD_BG = "D9F0D9"
REP_HDR = "7B3F00"
REP_BG = "FFF0CC"
MILESTONE_BG = "FFD966"
MILESTONE_FG = "7B3F00"
WARN_BG = "FFE0E0"
WARN_RED = "C00000"
TEAL = "006B6B"
TEAL_LT = "E0F0F0"
AM_CLR = "7030A0"
QC_CLR = "C55A11"
RE_CLR = "006B6B"
MS_BG = "1F3864"
WARN1_BG = "FFC000"
WARN2_BG = "FF9900"


# ── Resource tab layout constants ─────────────────────────
MAX_PLAN = 25
MAX_FIELD = 40
MAX_REP = 20
PLAN_DIV = 7
PLAN_S = 8
PLAN_E = PLAN_S + MAX_PLAN - 1
FIELD_DIV = 33
FIELD_S = 34
FIELD_E = FIELD_S + MAX_FIELD - 1
REP_DIV = 74
REP_S = 75
REP_E = REP_S + MAX_REP - 1
TOT_R = 95
VAR_R = 96
DEFAULT_PLAN = 4
DEFAULT_FIELD = 16
DEFAULT_REP = 4


# ── Audit Setup row references ────────────────────────────
AS_KICK = 5
AS_REL = 6
AS_HOL = 7
AS_BUF = 8
AS_PLAN = 12
AS_FIELD = 13
AS_REP = 14
AS_TOT = 16

MS_ROWS = {
    "kickoff": 20,
    "end_plan": 21,
    "outline": 22,
    "ontarget": 23,
    "writing": 24,
    "draft": 25,
    "exit": 26,
    "mgmt": 27,
    "final": 28,
    "exec_app": 29,
    "pm_am": 30,
    "editors": 31,
    "release": 32,
}


# ── Master Budget by Date (MBDD) layout ───────────────────
MBDD_PLAN_DIV = 5
MBDD_PLAN_S = 6
MBDD_PLAN_E = 6 + MAX_PLAN - 1
MBDD_PLAN_TOT = 31
MBDD_FIELD_DIV = 32
MBDD_FIELD_S = 33
MBDD_FIELD_E = 33 + MAX_FIELD - 1
MBDD_FIELD_TOT = 73
MBDD_REP_DIV = 74
MBDD_REP_S = 75
MBDD_REP_E = 75 + MAX_REP - 1
MBDD_REP_TOT = 95
MBDD_GRAND = 96


# ── Budget by Task (BBT) layout ───────────────────────────
BBT_PLAN_DIV = 3
BBT_PLAN_ROW = 4
BBT_PLAN_TOT = 5
BBT_FIELD_DIV = 6
BBT_FIELD_S = 7
BBT_FIELD_E = 21
BBT_FIELD_TOT = 22
BBT_FIELD_VAL = 23
BBT_REP_DIV = 24
BBT_REP1 = 25
BBT_REP2 = 26
BBT_REP_TOT = 27
BBT_CC_DIV = 28
BBT_CC_AM = 29
BBT_CC_QC = 30
BBT_CC_RE = 31
BBT_CC_MTG = 32
BBT_CC_WRAP = 33
BBT_CC_MGMT = 34
BBT_CC_TOT = 35
BBT_GRAND = 36


# ── Resource definitions ──────────────────────────────────
RESOURCES = [
    ("Resource PM", "PM", "1F3864"),
    ("Resource Asst PM", "Asst PM", "2E5090"),
    ("Resource Auditor 1", "Auditor 1", "1D5C2C"),
    ("Resource Auditor 2", "Auditor 2", "375623"),
    ("Resource Auditor 3", "Auditor 3", "4A7C59"),
    ("Resource Auditor 4", "Auditor 4", "2D6A4F"),
    ("Resource AM", "AM", "7030A0"),
    ("Resource QC", "QC", "C55A11"),
    ("Resource RE", "RE", "006B6B"),
]
RES_NAMES = [r[1] for r in RESOURCES]
RES_TABS = [r[0] for r in RESOURCES]
MBDD_RES_COLS = {name: 3 + i for i, name in enumerate(RES_NAMES)}
BBT_RES_COLS = MBDD_RES_COLS
STAFF_RES = RES_NAMES[:6]
OH_RES = RES_NAMES[6:]


# ── Helper functions for openpyxl style objects ──────────
def make_font(bold=False, color=BLACK, size=10, italic=False):
    """Create an Arial font with the given properties."""
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)


def make_fill(hex_color):
    """Create a solid fill with the given hex color."""
    return PatternFill("solid", start_color=hex_color, end_color=hex_color)


def make_align(horizontal="left", vertical="center", wrap=False):
    """Create an alignment with the given properties."""
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)


def make_border(style="thin", color="BFBFBF"):
    """Create a four-sided border with uniform style and color."""
    side = Side(style=style, color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def column_letter(index):
    """Return the letter for a 1-indexed column number."""
    return get_column_letter(index)
