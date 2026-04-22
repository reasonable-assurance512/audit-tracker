"""
Audit Resource Tracker v4 — reference workbook generator.

Reconstructed from PDF source for use as the Sprint 2 golden file.
Produces a 13-tab workbook: Holidays & Skeleton, Audit Setup,
9 Resource tabs, Master Budget by Date, Budget by Task.
(Audit Timeline tab is added separately by add_timeline.py.)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from datetime import date

wb = Workbook()

# ── Colors ──────────────────────────────────────────────────
DARK_BLUE="1F3864"; MED_BLUE="2E5090"; LIGHT_BLUE="DCE6F1"
DARK_RED="C00000"; LIGHT_RED="FFE0E0"
DARK_GRN="375623"; LIGHT_GRN="EBF1DE"
YELLOW="FFF2CC"; ORANGE="ED7D31"; BROWN="7B3F00"
GRAY_LT="F2F2F2"; GRAY_MD="D9D9D9"
WHITE="FFFFFF"; BLACK="000000"
BLUE_IN="0000CD"; RED_F="C00000"
PURPLE="7030A0"; LIGHT_PURPLE="E8DDEF"
PLAN_HDR="2C4770"; PLAN_BG="D6E4F0"
FIELD_HDR="1D5C2C"; FIELD_BG="D9F0D9"
REP_HDR="7B3F00"; REP_BG="FFF0CC"
MILESTONE_BG="FFD966"; MILESTONE_FG="7B3F00"
WARN_BG="FFE0E0"; WARN_RED="C00000"
TEAL="006B6B"; TEAL_LT="E0F0F0"
AM_CLR="7030A0"; QC_CLR="C55A11"; RE_CLR="006B6B"
MS_BG="1F3864"; WARN1_BG="FFC000"; WARN2_BG="FF9900"

def F(bold=False,color=BLACK,size=10,italic=False):
    return Font(name="Arial",bold=bold,color=color,size=size,italic=italic)
def Fill(h): return PatternFill("solid",start_color=h,end_color=h)
def Align(h="left",v="center",wrap=False):
    return Alignment(horizontal=h,vertical=v,wrap_text=wrap)
def Bdr(t="thin",c="BFBFBF"):
    s=Side(style=t,color=c); return Border(left=s,right=s,top=s,bottom=s)
def cl(idx): return get_column_letter(idx)

# ── Layout constants (resource tabs) ──────────────────────
MAX_PLAN=25; MAX_FIELD=40; MAX_REP=20
PLAN_DIV=7; PLAN_S=8; PLAN_E=PLAN_S+MAX_PLAN-1
FIELD_DIV=33; FIELD_S=34; FIELD_E=FIELD_S+MAX_FIELD-1
REP_DIV=74; REP_S=75; REP_E=REP_S+MAX_REP-1
TOT_R=95; VAR_R=96
DEFAULT_PLAN=4; DEFAULT_FIELD=16; DEFAULT_REP=4

# ── Audit Setup row references ─────────────────────────────
AS_KICK=5; AS_REL=6; AS_HOL=7; AS_BUF=8
AS_PLAN=12; AS_FIELD=13; AS_REP=14; AS_TOT=16

MS_ROWS = {
    'kickoff':20,'end_plan':21,'outline':22,'ontarget':23,
    'writing':24,'draft':25,'exit':26,'mgmt':27,
    'final':28,'exec_app':29,'pm_am':30,'editors':31,'release':32
}

# ── MBDD layout ────────────────────────────────────────────
MBDD_PLAN_DIV=5; MBDD_PLAN_S=6; MBDD_PLAN_E=6+MAX_PLAN-1
MBDD_PLAN_TOT=31
MBDD_FIELD_DIV=32; MBDD_FIELD_S=33; MBDD_FIELD_E=33+MAX_FIELD-1
MBDD_FIELD_TOT=73
MBDD_REP_DIV=74; MBDD_REP_S=75; MBDD_REP_E=75+MAX_REP-1
MBDD_REP_TOT=95; MBDD_GRAND=96

# ── BBT layout ─────────────────────────────────────────────
BBT_PLAN_DIV=3; BBT_PLAN_ROW=4; BBT_PLAN_TOT=5
BBT_FIELD_DIV=6; BBT_FIELD_S=7; BBT_FIELD_E=21
BBT_FIELD_TOT=22; BBT_FIELD_VAL=23
BBT_REP_DIV=24; BBT_REP1=25; BBT_REP2=26; BBT_REP_TOT=27
BBT_CC_DIV=28; BBT_CC_AM=29; BBT_CC_QC=30; BBT_CC_RE=31
BBT_CC_MTG=32; BBT_CC_WRAP=33; BBT_CC_MGMT=34
BBT_CC_TOT=35; BBT_GRAND=36

# ── Resource tabs ──────────────────────────────────────────
RESOURCES = [
    ("Resource PM", "PM", "1F3864"),
    ("Resource Asst PM", "Asst PM", "2E5090"),
    ("Resource Auditor 1","Auditor 1","1D5C2C"),
    ("Resource Auditor 2","Auditor 2","375623"),
    ("Resource Auditor 3","Auditor 3","4A7C59"),
    ("Resource Auditor 4","Auditor 4","2D6A4F"),
    ("Resource AM", "AM", "7030A0"),
    ("Resource QC", "QC", "C55A11"),
    ("Resource RE", "RE", "006B6B"),
]
RES_NAMES = [r[1] for r in RESOURCES]
RES_TABS = [r[0] for r in RESOURCES]

MBDD_RES_COLS = {name: 3+i for i,name in enumerate(RES_NAMES)}
BBT_RES_COLS = MBDD_RES_COLS
STAFF_RES = RES_NAMES[:6]
OH_RES = RES_NAMES[6:]

# ── Holidays / Skeleton data ───────────────────────────────
CLOSED=[
    ("Labor Day", date(2025, 9, 1),"FY2026"),
    ("Veterans Day", date(2025,11,11),"FY2026"),
    ("Thanksgiving Day", date(2025,11,27),"FY2026"),
    ("Day after Thanksgiving", date(2025,11,28),"FY2026"),
    ("Christmas Eve Day", date(2025,12,24),"FY2026"),
    ("Christmas Day", date(2025,12,25),"FY2026"),
    ("Day after Christmas", date(2025,12,26),"FY2026"),
    ("New Year's Day", date(2026, 1, 1),"FY2026"),
    ("Martin Luther King, Jr. Day", date(2026, 1,19),"FY2026"),
    ("Presidents' Day", date(2026, 2,16),"FY2026"),
    ("Memorial Day", date(2026, 5,25),"FY2026"),
    ("Labor Day", date(2026, 9, 7),"FY2027"),
    ("Veterans Day", date(2026,11,11),"FY2027"),
    ("Thanksgiving Day", date(2026,11,26),"FY2027"),
    ("Day after Thanksgiving", date(2026,11,27),"FY2027"),
    ("Christmas Eve Day", date(2026,12,24),"FY2027"),
    ("Christmas Day", date(2026,12,25),"FY2027"),
    ("New Year's Day", date(2027, 1, 1),"FY2027"),
    ("Martin Luther King, Jr. Day", date(2027, 1,18),"FY2027"),
    ("Presidents' Day", date(2027, 2,15),"FY2027"),
    ("Memorial Day", date(2027, 5,31),"FY2027"),
]
SKEL=[
    ("Texas Independence Day",date(2026, 3, 2),"FY2026"),
    ("San Jacinto Day", date(2026, 4,21),"FY2026"),
    ("Emancipation Day", date(2026, 6,19),"FY2026"),
    ("LBJ Day", date(2026, 8,27),"FY2026"),
    ("Confederate Heroes Day",date(2027, 1,19),"FY2027"),
    ("Texas Independence Day",date(2027, 3, 2),"FY2027"),
    ("San Jacinto Day", date(2027, 4,21),"FY2027"),
    ("LBJ Day", date(2027, 8,27),"FY2027"),
]

# ══════════════════════════════════════════════════════════
# TAB 1 — Holidays & Skeleton
# ══════════════════════════════════════════════════════════
hws=wb.active; hws.title="Holidays & Skeleton"
hws.sheet_properties.tabColor=DARK_BLUE
for col,w in [('A',30),('B',14),('C',14),('D',10)]:
    hws.column_dimensions[col].width=w

def write_hol_section(ws,start,title,hdr_bg,data):
    ws.merge_cells(f'A{start}:D{start}')
    ws[f'A{start}']=title
    ws[f'A{start}'].font=F(bold=True,color=WHITE,size=10)
    ws[f'A{start}'].fill=Fill(hdr_bg)
    ws[f'A{start}'].alignment=Align()
    ws.row_dimensions[start].height=18
    hr=start+1
    for ci,h in enumerate(["Holiday","Date","Day of Week","FY"],1):
        c=ws.cell(row=hr,column=ci,value=h)
        c.font=F(bold=True,size=9); c.fill=Fill(GRAY_MD)
        c.alignment=Align("center"); c.border=Bdr()
    ws.row_dimensions[hr].height=16
    ds=hr+1; r=ds
    for name,dt,fy in data:
        ws.cell(row=r,column=1,value=name).font=F(size=9)
        dc=ws.cell(row=r,column=2,value=dt)
        dc.number_format='MM/DD/YYYY'; dc.font=F(size=9)
        ws.cell(row=r,column=3,value=dt.strftime('%A')).font=F(size=9)
        ws.cell(row=r,column=4,value=fy).font=F(size=9)
        for ci in range(1,5):
            cell=ws.cell(row=r,column=ci)
            cell.border=Bdr()
            cell.alignment=Align("center" if ci>1 else "left")
        ws.row_dimensions[r].height=15
        r+=1
    return ds, r-1

hws.merge_cells('A1:D1')
hws['A1']="State Holiday Reference — FY2026 & FY2027 (Weekday Holidays Only)"
hws['A1'].font=F(bold=True,color=WHITE,size=12)
hws['A1'].fill=Fill(DARK_BLUE); hws['A1'].alignment=Align("center")
hws.row_dimensions[1].height=22

r_cs,r_ce=write_hol_section(hws,3,"ALL AGENCIES CLOSED",DARK_BLUE,CLOSED)
r_ss,r_se=write_hol_section(hws,r_ce+2,"SKELETON CREW DAYS",BROWN,SKEL)

nr=r_se+2
hws.merge_cells(f'A{nr}:D{nr}')
hws[f'A{nr}']="Weekend holidays excluded from this table."
hws[f'A{nr}'].font=F(italic=True,size=8,color="595959")
hws[f'A{nr}'].fill=Fill(GRAY_LT)
hws[f'A{nr}'].alignment=Align("left",wrap=True)
hws.row_dimensions[nr].height=28

CR=f"'Holidays & Skeleton'!$B${r_cs}:$B${r_ce}"
SR=f"'Holidays & Skeleton'!$B${r_ss}:$B${r_se}"
print(f"H&S: closed {r_cs}-{r_ce}, skel {r_ss}-{r_se}")

# ══════════════════════════════════════════════════════════
# TAB 2 — Audit Setup
# ══════════════════════════════════════════════════════════
sws=wb.create_sheet("Audit Setup")
sws.sheet_properties.tabColor=PURPLE
for col,w in [('A',36),('B',16),('C',16),('D',16),('E',46)]:
    sws.column_dimensions[col].width=w

def setup_hdr(ws,row,text,bg,span=5):
    ws.merge_cells(f'A{row}:{cl(span)}{row}')
    ws[f'A{row}']=text
    ws[f'A{row}'].font=F(bold=True,color=WHITE,size=10)
    ws[f'A{row}'].fill=Fill(bg)
    ws[f'A{row}'].alignment=Align()
    ws.row_dimensions[row].height=18

def hdr_row(ws,row,items,height=18):
    ws.row_dimensions[row].height=height
    for ci,(lbl,bg,fg) in enumerate(items,1):
        c=ws.cell(row=row,column=ci,value=lbl)
        c.font=F(bold=True,color=fg,size=9)
        c.fill=Fill(bg); c.alignment=Align("center"); c.border=Bdr()

def inp_row(ws,row,label,value,fmt=None,note="",note_bold=False):
    ws.row_dimensions[row].height=18
    c=ws.cell(row=row,column=1,value=label)
    c.font=F(bold=True,size=10); c.border=Bdr(); c.alignment=Align()
    v=ws.cell(row=row,column=2,value=value)
    v.font=F(bold=True,color=BLUE_IN,size=11)
    v.fill=Fill(LIGHT_BLUE); v.border=Bdr(); v.alignment=Align("center")
    if fmt: v.number_format=fmt
    if note:
        ws.merge_cells(f'C{row}:E{row}')
        nc=ws.cell(row=row,column=3,value=note)
        nc.font=F(bold=note_bold,size=9,color="595959" if not note_bold else BLACK)
        nc.alignment=Align("left",wrap=True); nc.border=Bdr()

sws.merge_cells('A1:E1')
sws['A1']="Audit Configuration & Milestone Calculator | v4"
sws['A1'].font=F(bold=True,color=WHITE,size=13)
sws['A1'].fill=Fill(PURPLE); sws['A1'].alignment=Align("center")
sws.row_dimensions[1].height=26
sws.row_dimensions[2].height=6

setup_hdr(sws,3,"ANCHOR DATES",PURPLE)
hdr_row(sws,4,[("Field",GRAY_MD,WHITE),("Value",GRAY_MD,WHITE),
    ("Day of Week / Calc",GRAY_MD,WHITE),("",GRAY_MD,WHITE),
    ("Validation / Notes",GRAY_MD,WHITE)])

sws.row_dimensions[5].height=18
sws.cell(row=5,column=1,value="Kickoff / Project Launch Date").font=F(bold=True,size=10)
sws.cell(row=5,column=1).border=Bdr()
sws.cell(row=5,column=1).alignment=Align()
kc=sws.cell(row=5,column=2,value=date(2026,4,1))
kc.font=F(bold=True,color=BLUE_IN,size=11); kc.fill=Fill(LIGHT_BLUE)
kc.number_format='MM/DD/YYYY'; kc.border=Bdr(); kc.alignment=Align("center")
sws.merge_cells('C5:D5')
sws.cell(row=5,column=3,value='=TEXT(B5,"dddd")').font=F(size=10)
sws.cell(row=5,column=3).alignment=Align("center")
sws.cell(row=5,column=3).border=Bdr()
sws.cell(row=5,column=5,value='=IF(WEEKDAY(B5,2)=5,"\u26a0 Kickoff is a Friday","\u2713 OK")')
sws.cell(row=5,column=5).alignment=Align()
sws.cell(row=5,column=5).border=Bdr()

sws.row_dimensions[6].height=18
sws.cell(row=6,column=1,value="Report Release Date (auto-calculated)").font=F(bold=True,size=10)
sws.cell(row=6,column=1).border=Bdr()
sws.cell(row=6,column=1).alignment=Align()
rc=sws.cell(row=6,column=2,
    value=f"=WORKDAY(B5-WEEKDAY(B5,2)+1+(B{AS_TOT}*7)-3,1,{CR})")
rc.font=F(bold=True,size=11); rc.fill=Fill(LIGHT_GRN)
rc.number_format='MM/DD/YYYY'; rc.border=Bdr(); rc.alignment=Align("center")
sws.merge_cells('C6:D6')
sws.cell(row=6,column=3,value='=TEXT(B6,"dddd")').font=F(size=10)
sws.cell(row=6,column=3).alignment=Align("center")
sws.cell(row=6,column=3).border=Bdr()
sws.cell(row=6,column=5,
    value='=IF(WEEKDAY(B6,2)<>1,"\u26a0 Release is not Monday","\u2713 OK")')
sws.cell(row=6,column=5).border=Bdr()
sws.cell(row=6,column=5).alignment=Align()

inp_row(sws,7,"Standard Hours Lost per Holiday/Skeleton Day",8,'0',
    "Deducted from resource tabs per closure. Default: 8.")

sws.row_dimensions[8].height=18
sws.cell(row=8,column=1,value="Weeks Before End of Fieldwork for On-Target Meeting").font=F(bold=True,size=10)
sws.cell(row=8,column=1).border=Bdr()
sws.cell(row=8,column=1).alignment=Align()
bc=sws.cell(row=8,column=2,value=1)
bc.font=F(bold=True,color=BLUE_IN,size=11); bc.fill=Fill(LIGHT_BLUE)
bc.number_format='0'; bc.border=Bdr(); bc.alignment=Align("center")
sws.merge_cells('C8:E8')
sws.cell(row=8,column=3,
    value='="On-Target = "&TEXT(D13-4-$B$8*7,"MM/DD/YY dddd")&" | Buffer: "&$B$8&" wk(s)"')
sws.cell(row=8,column=3).alignment=Align("left")
sws.cell(row=8,column=3).border=Bdr()

sws.row_dimensions[9].height=8

sws.conditional_formatting.add('E6',
    FormulaRule(formula=['=WEEKDAY($B$6,2)<>1'],
        fill=Fill(WARN_BG),font=Font(name="Arial",color=WARN_RED,bold=True,size=9)))

setup_hdr(sws,10,"PHASE CONFIGURATION",PLAN_HDR)
hdr_row(sws,11,[("Phase",GRAY_MD,WHITE),("Weeks Allocated",GRAY_MD,WHITE),
    ("Start (Mon)",GRAY_MD,WHITE),("End (Fri)",GRAY_MD,WHITE),
    ("Notes / Validation",GRAY_MD,WHITE)])

PHASES=[
    (AS_PLAN,"PLANNING", 4, PLAN_HDR,PLAN_BG,
        "=B5-WEEKDAY(B5,2)+1",f"=C{AS_PLAN}+(B{AS_PLAN}-1)*7+4",
        "Start snaps to Monday of kickoff week"),
    (AS_FIELD,"FIELDWORK",16,FIELD_HDR,FIELD_BG,
        f"=D{AS_PLAN}+3",f"=C{AS_FIELD}+(B{AS_FIELD}-1)*7+4",
        ""),
    (AS_REP,"REPORTING", 4, REP_HDR,REP_BG,
        f"=D{AS_FIELD}+3",f"=C{AS_REP}+(B{AS_REP}-1)*7+4",
        "Begins Monday after Fieldwork ends"),
]
for row,phase,wks,hc,bg,sf,ef,note in PHASES:
    sws.row_dimensions[row].height=20
    pc=sws.cell(row=row,column=1,value=phase)
    pc.font=F(bold=True,color=WHITE,size=11); pc.fill=Fill(hc)
    pc.alignment=Align(); pc.border=Bdr()
    wc=sws.cell(row=row,column=2,value=wks)
    wc.font=F(bold=True,color=BLUE_IN,size=12); wc.fill=Fill(LIGHT_BLUE)
    wc.alignment=Align("center"); wc.border=Bdr(); wc.number_format='0" wks"'
    sc=sws.cell(row=row,column=3,value=sf)
    sc.font=F(size=10); sc.number_format='MM/DD/YY'; sc.fill=Fill(bg)
    sc.alignment=Align("center"); sc.border=Bdr()
    ec=sws.cell(row=row,column=4,value=ef)
    ec.font=F(size=10); ec.number_format='MM/DD/YY'; ec.fill=Fill(bg)
    ec.alignment=Align("center"); ec.border=Bdr()
    if note:
        sws.cell(row=row,column=5,value=note).font=F(italic=True,size=9,color="595959")
        sws.cell(row=row,column=5).alignment=Align()
        sws.cell(row=row,column=5).border=Bdr()

fw_val=f'=IF(C{MS_ROWS["draft"]}-C{MS_ROWS["ontarget"]}<28,"\u26a0 Writing window short","\u2713 OK")'
sws.cell(row=AS_FIELD,column=5,value=fw_val).font=F(bold=True,size=9)
sws.cell(row=AS_FIELD,column=5).alignment=Align(wrap=True)
sws.cell(row=AS_FIELD,column=5).border=Bdr()
sws.row_dimensions[AS_FIELD].height=28

sws.conditional_formatting.add(f'E{AS_FIELD}',
    FormulaRule(formula=[f'=C{MS_ROWS["draft"]}-C{MS_ROWS["ontarget"]}<28'],
        fill=Fill(WARN_BG),font=Font(name="Arial",color=WARN_RED,bold=True,size=9)))

sws.row_dimensions[15].height=8
sws.row_dimensions[AS_TOT].height=18
sws.cell(row=AS_TOT,column=1,value="TOTAL AUDIT WEEKS").font=F(bold=True,size=10)
sws.cell(row=AS_TOT,column=1).border=Bdr()
sws.cell(row=AS_TOT,column=1).fill=Fill(GRAY_LT)
sws.cell(row=AS_TOT,column=1).alignment=Align()
tc=sws.cell(row=AS_TOT,column=2,value=f'=SUM(B{AS_PLAN}:B{AS_REP})')
tc.font=F(bold=True,size=12); tc.fill=Fill(GRAY_MD); tc.border=Bdr()
tc.alignment=Align("center"); tc.number_format='0" wks"'
sws.merge_cells(f'C{AS_TOT}:E{AS_TOT}')
sws.cell(row=AS_TOT,column=3,
    value=f'="Audit: "&TEXT(C{AS_PLAN},"MM/DD/YY")&" - "&TEXT(D{AS_REP},"MM/DD/YY")')
sws.cell(row=AS_TOT,column=3).alignment=Align()
sws.cell(row=AS_TOT,column=3).border=Bdr()

sws.row_dimensions[17].height=8

setup_hdr(sws,18,"MILESTONE CALENDAR",PURPLE)
hdr_row(sws,19,[("Milestone",GRAY_MD,WHITE),("",GRAY_MD,WHITE),
    ("Date",GRAY_MD,WHITE),("Day of Week",GRAY_MD,WHITE),
    ("Notes / Validation",GRAY_MD,WHITE)])

MILESTONES=[
    (MS_ROWS['kickoff'], "Kickoff / Project Launch",
        f"=B{AS_KICK}","Anchor",True),
    (MS_ROWS['end_plan'], "End of Planning",
        f"=D{AS_PLAN}","Last Friday of Planning",False),
    (MS_ROWS['outline'], "Outline Sent to Exec",
        f"=WORKDAY(C{MS_ROWS['ontarget']},-2,{CR})","2 business days before On-Target",False),
    (MS_ROWS['ontarget'], "ON-TARGET MEETING",
        f"=D{AS_FIELD}-4-$B$8*7","Monday of week B8 weeks before end of Fieldwork",True),
    (MS_ROWS['writing'], "Writing Period Begins",
        f"=C{MS_ROWS['ontarget']}+1","Day after On-Target",False),
    (MS_ROWS['draft'], "Draft Sent to Exec",
        f"=WORKDAY(C{MS_ROWS['exit']},-3,{CR})","3 business days before Exit Conference",False),
    (MS_ROWS['exit'], "Exit Conference",
        f"=WORKDAY(C{MS_ROWS['mgmt']},-10,{CR})","10 business days before Mgmt Responses Due",False),
    (MS_ROWS['mgmt'], "Management Responses Due",
        f"=B{AS_REL}-7","Monday of release week minus 1 week",False),
    (MS_ROWS['final'], "Final Draft to Exec",
        f"=B{AS_REL}-6","Tuesday of release week",False),
    (MS_ROWS['exec_app'], "Exec Approves Final Draft",
        f"=B{AS_REL}-5","Wednesday of release week",False),
    (MS_ROWS['pm_am'], "PM / AM Review Prototype",
        f"=B{AS_REL}-5","Wednesday-Thursday of release week",False),
    (MS_ROWS['editors'], "Editors Set Up Release Emails",
        f"=B{AS_REL}-3","Friday of release week",False),
    (MS_ROWS['release'], "REPORT RELEASE",
        f"=B{AS_REL}","Anchor",True),
]
for row,name,formula,desc,anchor in MILESTONES:
    sws.row_dimensions[row].height=18
    bg=LIGHT_BLUE if anchor else WHITE
    nc=sws.cell(row=row,column=1,value=name)
    nc.font=F(bold=anchor,size=10 if anchor else 9)
    nc.fill=Fill(bg); nc.alignment=Align(); nc.border=Bdr()
    sws.cell(row=row,column=2).border=Bdr()
    dc=sws.cell(row=row,column=3,value=formula)
    dc.font=F(bold=anchor,size=10); dc.number_format='MM/DD/YYYY'
    dc.fill=Fill(bg); dc.alignment=Align("center"); dc.border=Bdr()
    dow=sws.cell(row=row,column=4,value=f'=IF(ISNUMBER(C{row}),TEXT(C{row},"dddd"),"")')
    dow.font=F(size=9); dow.fill=Fill(bg); dow.alignment=Align("center"); dow.border=Bdr()
    ec=sws.cell(row=row,column=5,value=desc)
    ec.font=F(italic=True,size=9,color="595959")
    ec.fill=Fill(bg); ec.alignment=Align(); ec.border=Bdr()

sws.cell(row=MS_ROWS['draft'],column=5,
    value=f'=IF(C{MS_ROWS["draft"]}-C{MS_ROWS["ontarget"]}<28,"\u26a0 Writing window short","")')
sws.cell(row=MS_ROWS['draft'],column=5).border=Bdr()
sws.row_dimensions[MS_ROWS['draft']].height=24

sws.cell(row=MS_ROWS['release'],column=5,
    value=f'=IF(WEEKDAY(B{AS_REL},2)<>1,"\u26a0 Release not Monday","\u2713 OK")')
sws.cell(row=MS_ROWS['release'],column=5).border=Bdr()

sws.conditional_formatting.add(f'A{MS_ROWS["draft"]}:E{MS_ROWS["draft"]}',
    FormulaRule(formula=[f'=C{MS_ROWS["draft"]}-C{MS_ROWS["ontarget"]}<28'],
        fill=Fill(WARN_BG),font=Font(name="Arial",color=WARN_RED,bold=True,size=9)))

sws.conditional_formatting.add(f'A{MS_ROWS["release"]}:E{MS_ROWS["release"]}',
    FormulaRule(formula=[f'=WEEKDAY($B${AS_REL},2)<>1'],
        fill=Fill(WARN_BG),font=Font(name="Arial",color=WARN_RED,bold=True,size=9)))

sws.freeze_panes='A5'

# ══════════════════════════════════════════════════════════
# RESOURCE TAB BUILDER (used 9 times)
# ══════════════════════════════════════════════════════════
RES_COLS = {
    'A':11,'B':24,'C':14,'D':16,'E':16,
    'F':16,'G':16,'H':16,'I':4,'J':4,'K':16,'L':36
}

def milestone_col_f(c_ref):
    checks=[
        (MS_ROWS['kickoff'], "\u2691 KICKOFF"),
        (MS_ROWS['end_plan'], "\u2691 END OF PLANNING"),
        (MS_ROWS['outline'], "\u2691 OUTLINE TO EXEC"),
        (MS_ROWS['ontarget'], "\u2691 ON-TARGET"),
        (MS_ROWS['draft'], "\u2691 DRAFT TO EXEC"),
        (MS_ROWS['exit'], "\u2691 EXIT CONFERENCE"),
        (MS_ROWS['mgmt'], "\u2691 MGMT RESPONSES DUE"),
        (MS_ROWS['release'], "\u2691 REPORT RELEASE"),
    ]
    parts=[f'IF(AND(\'Audit Setup\'!$C${r}>={c_ref},\'Audit Setup\'!$C${r}<={c_ref}+4),"{lbl} ","")' for r,lbl in checks]
    inner = ",".join(parts)
    return '=IF(NOT(ISNUMBER(' + c_ref + ')),"",TRIM(CONCATENATE(' + inner + ')))'

def build_resource_tab(wb, tab_name, display_name, tab_color):
    ws=wb.create_sheet(tab_name)
    ws.sheet_properties.tabColor=tab_color
    for col,w in RES_COLS.items():
        ws.column_dimensions[col].width=w

    ws.merge_cells('A1:L1')
    ws['A1']=f"Audit Resource Tracker | v4 | {display_name}"
    ws['A1'].font=F(bold=True,color=WHITE,size=12)
    ws['A1'].fill=Fill(DARK_BLUE); ws['A1'].alignment=Align("center")
    ws.row_dimensions[1].height=24

    for row,lbl in [(2,"Resource Name:"),(3,"Audit / Project:")]:
        ws.row_dimensions[row].height=18
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}']=lbl
        ws[f'A{row}'].font=F(bold=True,size=10)
        ws[f'A{row}'].alignment=Align()
        ws.merge_cells(f'C{row}:L{row}')
        ws[f'C{row}']=display_name if row==2 else "Enter audit name"
        ws[f'C{row}'].font=F(color=BLUE_IN,size=10)
        ws[f'C{row}'].fill=Fill(LIGHT_BLUE)
        ws[f'C{row}'].alignment=Align()

    ws.merge_cells('A4:L4')
    ws['A4']="Phase structure driven by 'Audit Setup'. Blue = manual entry."
    ws['A4'].font=F(italic=True,size=8,color="595959")
    ws['A4'].fill=Fill(GRAY_LT); ws['A4'].alignment=Align("left",wrap=True)
    ws.row_dimensions[4].height=20
    ws.row_dimensions[5].height=6

    ws.row_dimensions[6].height=42
    HDRS=[
        (1,"Week\n#", DARK_BLUE,WHITE),
        (2,"Task\nAssignment",PLAN_HDR,WHITE),
        (3,"Week Of\n(Mon)", DARK_BLUE,WHITE),
        (4,"Hours\nAvail.", PLAN_HDR,WHITE),
        (5,"(-) Business\nHolidays", DARK_RED,WHITE),
        (6,"(-) Skeleton\nDays", BROWN,WHITE),
        (7,"(-) Leave\nPlanned", DARK_RED,WHITE),
        (8,"Total\nAvail.", DARK_GRN,WHITE),
        (9,"",WHITE,WHITE),(10,"",WHITE,WHITE),
        (11,"Actual Hrs", DARK_BLUE,WHITE),
        (12,"Milestones\nThis Week", PURPLE,WHITE),
    ]
    for ci,lbl,bg,fg in HDRS:
        c=ws.cell(row=6,column=ci,value=lbl)
        c.font=F(bold=True,color=fg,size=9)
        c.fill=Fill(bg); c.alignment=Align("center",wrap=True)
        if ci not in (9,10): c.border=Bdr()

    def phase_div(row,formula,bg):
        ws.row_dimensions[row].height=22
        ws.merge_cells(f'A{row}:L{row}')
        ws[f'A{row}']=formula
        ws[f'A{row}'].font=F(bold=True,color=WHITE,size=11)
        ws[f'A{row}'].fill=Fill(bg); ws[f'A{row}'].alignment=Align("left")

    phase_div(PLAN_DIV,
        f'=" PLANNING - Weeks 1-"&\'Audit Setup\'!$B${AS_PLAN}',
        PLAN_HDR)
    phase_div(FIELD_DIV,
        f'=" FIELDWORK - Weeks "&(\'Audit Setup\'!$B${AS_PLAN}+1)&"-"&(\'Audit Setup\'!$B${AS_PLAN}+\'Audit Setup\'!$B${AS_FIELD})',
        FIELD_HDR)
    phase_div(REP_DIV,
        f'=" REPORTING - Weeks "&(\'Audit Setup\'!$B${AS_PLAN}+\'Audit Setup\'!$B${AS_FIELD}+1)&"-"&\'Audit Setup\'!$B${AS_TOT}',
        REP_HDR)

    def write_rows(rows, phase_bg, max_cell, prefix, first_c_formula):
        for i,r in enumerate(rows):
            pos=i+1
            ws.row_dimensions[r].height=15
            ca=ws.cell(row=r,column=1,
                value=f"=IF({pos}<='Audit Setup'!{max_cell},\"{prefix}-\"&{pos},\"\")")
            ca.font=F(bold=True,size=9); ca.fill=Fill(phase_bg)
            ca.alignment=Align("center"); ca.border=Bdr()
            cb=ws.cell(row=r,column=2,value="")
            cb.font=F(size=9,italic=True,color="595959")
            cb.fill=Fill(LIGHT_BLUE); cb.alignment=Align("left")
            cb.border=Bdr()
            if i==0: c_val=first_c_formula
            else: c_val=f"=C{r-1}+7"
            cc=ws.cell(row=r,column=3,value=c_val)
            cc.number_format='MM/DD/YY'; cc.font=F(size=9)
            cc.fill=Fill(phase_bg); cc.alignment=Align("center"); cc.border=Bdr()
            cd=ws.cell(row=r,column=4,value=0)
            cd.font=F(color=BLUE_IN,size=9); cd.fill=Fill(LIGHT_BLUE)
            cd.number_format='#,##0.0;[Red]-#,##0.0;"-"'
            cd.alignment=Align("center"); cd.border=Bdr()
            ce=ws.cell(row=r,column=5,
                value=f"=-COUNTIFS({CR},\">=\"&C{r},{CR},\"<=\"&(C{r}+4))*'Audit Setup'!$B${AS_HOL}")
            ce.font=F(color=RED_F,size=9); ce.fill=Fill(phase_bg)
            ce.number_format='#,##0.0;-#,##0.0;"-"'
            ce.alignment=Align("center"); ce.border=Bdr()
            cf=ws.cell(row=r,column=6,
                value=f"=-COUNTIFS({SR},\">=\"&C{r},{SR},\"<=\"&(C{r}+4))*'Audit Setup'!$B${AS_HOL}")
            cf.font=F(color=BROWN,size=9); cf.fill=Fill(phase_bg)
            cf.number_format='#,##0.0;-#,##0.0;"-"'
            cf.alignment=Align("center"); cf.border=Bdr()
            cg=ws.cell(row=r,column=7,value=0)
            cg.font=F(color=RED_F,size=9); cg.fill=Fill(LIGHT_BLUE)
            cg.number_format='#,##0.0;[Red]-#,##0.0;"-"'
            cg.alignment=Align("center"); cg.border=Bdr()
            ch=ws.cell(row=r,column=8,
                value=f"=CEILING(MAX(0,SUM(D{r}:G{r})),1)")
            ch.font=F(bold=True,size=9); ch.fill=Fill(LIGHT_GRN)
            ch.number_format='#,##0'; ch.alignment=Align("center"); ch.border=Bdr()
            for ci in (9,10):
                ws.cell(row=r,column=ci).fill=Fill(GRAY_LT)
            ck=ws.cell(row=r,column=11,value=0)
            ck.font=F(color=BLUE_IN,size=9); ck.fill=Fill(LIGHT_BLUE)
            ck.number_format='#,##0.0;-#,##0.0;"-"'
            ck.alignment=Align("center"); ck.border=Bdr()
            cl_cell=ws.cell(row=r,column=12,value=milestone_col_f(f"C{r}"))
            cl_cell.font=F(bold=True,color=MILESTONE_FG,size=9)
            cl_cell.fill=Fill(phase_bg); cl_cell.alignment=Align("left",wrap=True)
            cl_cell.border=Bdr()

    write_rows(list(range(PLAN_S,PLAN_E+1)),PLAN_BG,f'$B${AS_PLAN}',"PL",
        f"='Audit Setup'!$C${AS_PLAN}")
    write_rows(list(range(FIELD_S,FIELD_E+1)),FIELD_BG,f'$B${AS_FIELD}',"FW",
        f"='Audit Setup'!$C${AS_FIELD}")
    write_rows(list(range(REP_S,REP_E+1)),REP_BG,f'$B${AS_REP}',"RP",
        f"='Audit Setup'!$C${AS_REP}")

    all_data=f'A{PLAN_S}:L{REP_E}'
    ws.conditional_formatting.add(all_data,
        FormulaRule(formula=[f'=AND(LEN($L{PLAN_S})>0,ISNUMBER(SEARCH("\u2691",$L{PLAN_S})))'],
            fill=Fill(MILESTONE_BG),
            font=Font(name="Arial",bold=True,color=MILESTONE_FG,size=9)))

    dv=DataValidation(type="decimal",operator="between",formula1="0",formula2="40",
        error="0-40 hrs only",errorTitle="Invalid Entry",showErrorMessage=True)
    ws.add_data_validation(dv); dv.sqref=f"D{PLAN_S}:D{REP_E}"

    ws.row_dimensions[TOT_R].height=22
    ws.merge_cells(f'A{TOT_R}:C{TOT_R}')
    tc=ws.cell(row=TOT_R,column=1,value="TOTALS")
    tc.font=F(bold=True,color=WHITE,size=11); tc.fill=Fill(DARK_BLUE)
    tc.alignment=Align("center"); tc.border=Bdr()

    def srange(c): return f"SUM({c}{PLAN_S}:{c}{PLAN_E},{c}{FIELD_S}:{c}{FIELD_E},{c}{REP_S}:{c}{REP_E})"

    for ci,fc,bg in [(4,BLUE_IN,LIGHT_BLUE),(5,RED_F,LIGHT_RED),(6,BROWN,YELLOW),
        (7,RED_F,LIGHT_RED),(8,DARK_GRN,LIGHT_GRN),(11,BLUE_IN,LIGHT_BLUE)]:
        col=get_column_letter(ci)
        c=ws.cell(row=TOT_R,column=ci,value=f"={srange(col)}")
        c.font=F(bold=True,color=fc,size=11); c.fill=Fill(bg)
        c.number_format='#,##0'; c.alignment=Align("center"); c.border=Bdr()
    for ci in (9,10):
        ws.cell(row=TOT_R,column=ci).fill=Fill(GRAY_LT)
    ws.cell(row=TOT_R,column=12).fill=Fill(GRAY_LT)

    ws.row_dimensions[VAR_R].height=18
    ws.merge_cells(f'A{VAR_R}:G{VAR_R}')
    ws[f'A{VAR_R}']="Variance: Actual (K) minus Total Available (H)"
    ws[f'A{VAR_R}'].font=F(bold=True,italic=True,size=9)
    ws[f'A{VAR_R}'].fill=Fill(GRAY_LT)
    ws[f'A{VAR_R}'].alignment=Align("right")
    vc=ws.cell(row=VAR_R,column=8,value=f'=K{TOT_R}-H{TOT_R}')
    vc.font=F(bold=True,size=11); vc.number_format='#,##0;[Red]-#,##0;"-"'
    vc.alignment=Align("center"); vc.border=Bdr(); vc.fill=Fill(GRAY_LT)
    ws.merge_cells(f'I{VAR_R}:L{VAR_R}')
    ws[f'I{VAR_R}']="Positive = over budget | Negative = under budget"
    ws[f'I{VAR_R}'].font=F(italic=True,size=8,color="595959")
    ws[f'I{VAR_R}'].fill=Fill(GRAY_LT)
    ws[f'I{VAR_R}'].alignment=Align("left")

    ws.freeze_panes='A7'

    for i,r in enumerate(range(PLAN_S,PLAN_E+1)):
        if i>=DEFAULT_PLAN: ws.row_dimensions[r].hidden=True
    for i,r in enumerate(range(FIELD_S,FIELD_E+1)):
        if i>=DEFAULT_FIELD: ws.row_dimensions[r].hidden=True
    for i,r in enumerate(range(REP_S,REP_E+1)):
        if i>=DEFAULT_REP: ws.row_dimensions[r].hidden=True

    return ws

res_ws = {}
for tab_name, display_name, tab_color in RESOURCES:
    res_ws[display_name] = build_resource_tab(wb, tab_name, display_name, tab_color)
    print(f"  Built: {tab_name}")

out='Audit_Resource_Tracker_v4.xlsx'
wb.save(out)
print(f"\nSaved: {out}")
print(f"Tabs: {[ws.title for ws in wb.worksheets]}")
